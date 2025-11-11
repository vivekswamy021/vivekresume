import streamlit as st
import os
import pdfplumber
import docx
import openpyxl
import json
import tempfile
from groq import Groq
from gtts import gTTS 
import traceback
import re 
from dotenv import load_dotenv 
from datetime import date 
import csv 

# Ensure that UploadedFile class is accessible for type checking
from streamlit.runtime.uploaded_file_manager import UploadedFile

# -------------------------
# CONFIGURATION & API SETUP
# -------------------------

GROQ_MODEL = "llama-3.1-8b-instant"

# Options for LLM functions
section_options = ["name", "email", "phone", "skills", "education", "experience", "certifications", "projects", "strength", "personal_details", "github", "linkedin", "full resume"]
question_section_options = ["skills","experience", "certifications", "projects", "education"] 

# Default Categories for JD Filtering (New)
DEFAULT_JOB_TYPES = ["Full-time", "Contract", "Internship", "Remote", "Part-time"]
DEFAULT_ROLES = ["Software Engineer", "Data Scientist", "Product Manager", "HR Manager", "Marketing Specialist", "Operations Analyst"]

# Load environment variables from .env file
load_dotenv()

# Ensure GROQ_API_KEY is defined in your environment or .env file
GROQ_API_KEY = os.getenv('GROQ_API_KEY')

if not GROQ_API_KEY:
    # Use st.warning instead of st.error/st.stop to allow the script to run 
    # if the user is running it without a key, but alert them to the dependency.
    st.warning(
        "🚨 WARNING: GROQ_API_KEY environment variable not set. "
        "AI functionality (Parsing, Matching, Q&A) will not work. "
        "Please ensure a '.env' file exists with your key."
    )
    # Initialize a mock client to prevent immediate crash if key is missing, 
    # although all LLM functions will fail.
    class MockGroqClient:
        def chat(self):
            class Completions:
                def create(self, **kwargs):
                    raise ValueError("GROQ_API_KEY not set. AI functions disabled.")
            return Completions()
    
    client = MockGroqClient()
else:
    # Initialize Groq Client
    client = Groq(api_key=GROQ_API_KEY)


# -------------------------
# Utility: Navigation Manager
# -------------------------
def go_to(page_name):
    """Changes the current page in Streamlit's session state."""
    st.session_state.page = page_name

def clear_interview_state():
    """Clears all generated questions, answers, and the evaluation report."""
    st.session_state.interview_qa = []
    st.session_state.iq_output = ""
    st.session_state.evaluation_report = ""
    st.toast("Practice answers cleared.")
    # -------------------------
# CORE LOGIC: FILE HANDLING AND EXTRACTION
# -------------------------

def get_file_type(file_path):
    """Identifies the file type based on its extension."""
    ext = os.path.splitext(file_path)[1].lower().strip('.')
    
    # Mapping extensions to general types (md, csv, json are handled like txt extraction logic below)
    if ext == 'pdf':
        return 'pdf'
    elif ext == 'docx':
        return 'docx'
    elif ext == 'xlsx':
        return 'xlsx'
    elif ext in ['txt', 'json', 'md', 'markdown', 'csv', 'rtf']: # Grouping text-based and simple structured formats
        return ext
    else:
        return 'txt' 

def extract_content(file_type, file_path):
    """Extracts text content from various file types."""
    text = ''
    try:
        if file_type == 'pdf':
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + '\n'
        
        elif file_type == 'docx':
            doc = docx.Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
        
        elif file_type == 'xlsx':
            workbook = openpyxl.load_workbook(file_path)
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                text += f"--- Sheet: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    # Concatenate non-None cell values in the row
                    row_text = ' | '.join([str(c) for c in row if c is not None])
                    if row_text.strip():
                        text += row_text + '\n'
                text += "\n"
        
        elif file_type == 'csv':
             with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    text += ' | '.join(row) + '\n'
        
        # Handles txt, json, md, markdown, rtf, and the default case
        elif file_type in ['txt', 'json', 'md', 'markdown', 'csv', 'rtf'] or file_type not in ['pdf', 'docx', 'xlsx']:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()

        # Final check
        if not text.strip():
            return f"Error: {file_type.upper()} content extraction failed. The file appears empty or non-text content could not be read."
        
        return text
    
    except Exception as e:
        return f"Fatal Extraction Error: Failed to read file content ({file_type}). Error details: {e}"

# -------------------------
# LLM & Extraction Functions
# -------------------------

@st.cache_data(show_spinner="Extracting JD metadata...")
def extract_jd_metadata(jd_text):
    """Extracts structured metadata (Role, Job Type, Key Skills) from raw JD text."""
    if not GROQ_API_KEY:
        return {"role": "N/A", "job_type": "N/A", "key_skills": []}

    prompt = f"""Analyze the following Job Description and extract the key metadata.
    
    Job Description:
    {jd_text}
    
    Provide the output strictly as a JSON object with the following three keys:
    1.  **role**: The main job title (e.g., 'Data Scientist', 'Senior Software Engineer'). If not clear, default to 'General Analyst'.
    2.  **job_type**: The employment type (e.g., 'Full-time', 'Contract', 'Internship', 'Remote'). If not clear, default to 'Full-time'.
    3.  **key_skills**: A list of 5 to 10 most critical hard and soft skills required (e.g., ['Python', 'AWS', 'Teamwork', 'SQL']).
    
    Example Output: {{"role": "Software Engineer", "job_type": "Full-time", "key_skills": ["Python", "JavaScript", "React", "AWS", "Agile"]}}
    """
    content = ""
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        content = response.choices[0].message.content.strip()

        # Robust JSON extraction using regex
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        
        if json_match:
            json_str = json_match.group(0).strip()
            
            # Additional cleanup for fences, although regex should handle it
            if json_str.startswith('```json'):
                json_str = json_str[len('```json'):]
            if json_str.endswith('```'):
                json_str = json_str[:-len('```')]
            json_str = json_str.strip()
            
            parsed = json.loads(json_str)
        else:
            raise json.JSONDecodeError("Could not isolate a valid JSON structure from LLM response.", content, 0)
        
        # Ensure the output structure is always correct
        return {
            "role": parsed.get("role", "General Analyst"),
            "job_type": parsed.get("job_type", "Full-time"),
            "key_skills": [s.strip() for s in parsed.get("key_skills", []) if isinstance(s, str)]
        }

    except Exception:
        # Fallback in case of API error or malformed JSON
        return {"role": "General Analyst (LLM Error)", "job_type": "Full-time (LLM Error)", "key_skills": ["LLM Error", "Fallback"]}


@st.cache_data(show_spinner="Analyzing content with Groq LLM...")
def parse_with_llm(text, return_type='json'):
    """Sends resume text to the LLM for structured information extraction."""
    if text.startswith("Error"):
        return {"error": text, "raw_output": ""}
    if not GROQ_API_KEY:
        return {"error": "GROQ_API_KEY not set. Cannot run LLM parsing.", "raw_output": ""}

    prompt = f"""Extract the following information from the resume in structured JSON.
    Ensure all relevant details for each category are captured.
    - Name, - Email, - Phone, - Skills, - Education (list of degrees/institutions/dates), 
    - Experience (list of job roles/companies/dates/responsibilities), - Certifications (list), 
    - Projects (list of project names/descriptions/technologies), - Strength (list of personal strengths/qualities), 
    - Personal Details (e.g., address, date of birth, nationality), - Github (URL), - LinkedIn (URL)
    
    Resume Text:
    {text}
    
    Provide the output strictly as a JSON object.
    """
    content = ""
    parsed = {}
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2
        )
        content = response.choices[0].message.content.strip()

        # --- CRITICAL FIX: AGGRESSIVE JSON ISOLATION USING REGEX ---
        
        # 1. Attempt to find the full JSON object using regex (non-greedy from first '{' to last '}')
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        
        json_str = ""
        if json_match:
            json_str = json_match.group(0).strip()
            
            # Strip markdown fences if they still exist after the regex match
            # This is key to fixing the malformed JSON issue caused by surrounding text
            if json_str.startswith('```json'):
                json_str = json_str[len('```json'):]
            if json_str.endswith('```'):
                json_str = json_str[:-len('```')]
            
            # Clean up the string content further (especially for the 'Extra data' error)
            json_str = json_str.strip()
            
            # 2. Attempt to load the JSON
            parsed = json.loads(json_str)
        else:
            # If we can't find a clear JSON object using regex, raise an error
            raise json.JSONDecodeError("Could not isolate a valid JSON structure from LLM response.", content, 0)
        
        # --- END CRITICAL FIX ---

    except json.JSONDecodeError as e:
        # Include the raw content and the traceback for detailed debugging
        error_msg = f"JSON decoding error from LLM. LLM returned malformed JSON. Error: {e} | Malformed string segment:\n---\n{json_str[:200]}..."
        parsed = {"error": error_msg, "raw_output": content}
    except ValueError as e: # Catch the MockGroqClient error
        parsed = {"error": str(e), "raw_output": "AI functions disabled."}
    except Exception as e:
        error_msg = f"LLM API interaction error: {e}"
        parsed = {"error": error_msg, "raw_output": "No LLM response due to API error."}

    if return_type == 'json':
        return parsed
    elif return_type == 'markdown':
        if "error" in parsed:
            # Provide better formatting for the error output
            return f"**Error:** {parsed.get('error', 'Unknown parsing error')}\n\n**Raw Output (for debugging):**\n```\n{parsed.get('raw_output','')}\n```"
        
        md = ""
        for k, v in parsed.items():
            if v:
                md += f"**{k.replace('_', ' ').title()}**:\n"
                if isinstance(v, list):
                    for item in v:
                        if item: 
                            md += f"- {item}\n"
                elif isinstance(v, dict):
                    for sub_k, sub_v in v.items():
                        if sub_v:
                            md += f"  - {sub_k.replace('_', ' ').title()}: {sub_v}\n"
                    md += "\n"
                else:
                    md += f"  {v}\n"
                md += "\n"
        return md
    return {"error": "Invalid return_type"}


def extract_jd_from_linkedin_url(url: str) -> str:
    """
    Simulates JD content extraction from a LinkedIn URL.
    """
    try:
        job_title = "Data Scientist"
        try:
            # More robust title extraction from a common LinkedIn job URL format
            match = re.search(r'/jobs/view/([^/]+)', url) or re.search(r'/jobs/(\w+)', url)
            if match:
                job_title = match.group(1).split('?')[0].replace('-', ' ').title()
                if job_title.lower().startswith('view'): job_title = 'Data Scientist' # Fallback
        except:
            pass

        if "linkedin.com/jobs/" not in url:
             return f"[Error: Not a valid LinkedIn Job URL format: {url}]"

        
        # Simulated synthesized JD content 
        jd_text = f"""
        --- Simulated JD for: {job_title} ---
        
        **Company:** Quantum Analytics Inc.
        **Role:** {job_title}
        
        **Responsibilities:**
        - Develop and implement machine learning models to solve complex business problems.
        - Clean, transform, and analyze large datasets using Python/R and SQL.
        - Collaborate with engineering teams to deploy models into production environments.
        - Communicate findings and model performance to non-technical stakeholders.
        
        **Requirements:**
        - MS/PhD in Computer Science, Statistics, or a quantitative field.
        - 3+ years of experience as a Data Scientist.
        - Expertise in Python (Pandas, Scikit-learn, TensorFlow/PyTorch).
        - Experience with cloud platforms (AWS, Azure, or GCP).
        
        --- End Simulated JD ---
        """
        
        return jd_text.strip()
            
    except Exception as e:
        return f"[Fatal Extraction Error: Simulation failed for URL {url}. Error: {e}]"


def evaluate_jd_fit(job_description, parsed_json):
    """Evaluates how well a resume fits a given job description, including section-wise scores."""
    if not GROQ_API_KEY:
        return "AI Evaluation Disabled: GROQ_API_KEY not set."
    if not job_description.strip(): return "Please paste a job description."
    if "error" in parsed_json: return "Cannot evaluate due to resume parsing errors."
    
    relevant_resume_data = {
        'Skills': parsed_json.get('skills', 'Not found or empty'),
        'Experience': parsed_json.get('experience', 'Not found or empty'),
        'Education': parsed_json.get('education', 'Not found or empty'),
    }
    resume_summary = json.dumps(relevant_resume_data, indent=2)

    prompt = f"""Evaluate how well the following resume content matches the provided job description.
    
    Job Description: {job_description}
    
    Resume Sections for Analysis:
    {resume_summary}
    
    Provide a detailed evaluation structured as follows:
    1.  **Overall Fit Score:** A score out of 10.
    2.  **Section Match Percentages:** A percentage score for the match in the key sections (Skills, Experience, Education).
    3.  **Strengths/Matches:** Key points where the resume aligns well with the JD.
    4.  **Gaps/Areas for Improvement:** Key requirements in the JD that are missing or weak in the resume.
    5.  **Overall Summary:** A concise summary of the fit.
    
    **Format the output strictly as follows, ensuring the scores are easily parsable (use brackets or no brackets around scores):**
    Overall Fit Score: [Score]/10
    
    --- Section Match Analysis ---
    Skills Match: [XX]%
    Experience Match: [YY]%
    Education Match: [ZZ]%
    
    Strengths/Matches:
    - Point 1
    - Point 2
    
    Gaps/Areas for Improvement:
    - Point 1
    - Point 2
    
    Overall Summary: [Concise summary]
    """

    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.3
    )
    return response.choices[0].message.content.strip()


def evaluate_interview_answers(qa_list, parsed_json):
    """Evaluates the user's answers against the resume content and provides feedback."""
    if not GROQ_API_KEY:
        return "AI Evaluation Disabled: GROQ_API_KEY not set."
    if "error" in parsed_json: return "Cannot evaluate due to resume parsing errors."

    
    resume_summary = json.dumps(parsed_json, indent=2)
    
    qa_summary = "\n---\n".join([
        f"Q: {item['question']}\nA: {item['answer']}" 
        for item in qa_list
    ])
    
    prompt = f"""You are an expert HR Interviewer. Evaluate the candidate's answers based on the following:
    1.  **The Candidate's Resume Content (for context):**
        {resume_summary}
    2.  **The Candidate's Questions and Answers:**
        {qa_summary}

    For each Question-Answer pair, provide a score (out of 10) and detailed feedback. The feedback must include:
    * **Clarity & Accuracy:** How well the answer directly and accurately addresses the question, referencing the resume context.
    * **Gaps & Improvements:** Specific suggestions on how the candidate could improve the answer or what critical resume points they missed/could elaborate on.
    
    Finally, provide an **Overall Summary** and a **Total Score (out of {len(qa_list) * 10})**.
    
    **Format the output strictly using Markdown headings and bullet points:**
    
    ## Evaluation Results
    
    ### Question 1: [Question Text]
    Score: [X]/10
    Feedback:
    - **Clarity & Accuracy:** ...
    - **Gaps & Improvements:** ...
    
    ### Question 2: [Question Text]
    Score: [X]/10
    Feedback:
    - **Clarity & Accuracy:** ...
    - **Gaps & Improvements:** ...
    
    ... [Repeat for all questions] ...
    
    ---
    
    ## Final Assessment
    Total Score: [Y]/{len(qa_list) * 10}
    Overall Summary: [A concise summary of the candidate's performance and next steps.]
    """

    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.3
    )
    return response.choices[0].message.content.strip()


def generate_interview_questions(parsed_json, section):
    """Generates categorized interview questions using LLM."""
    if not GROQ_API_KEY:
        return "AI Functions Disabled: GROQ_API_KEY not set."
    if "error" in parsed_json: return "Cannot generate questions due to resume parsing errors."
    
    section_title = section.replace("_", " ").title()
    section_content = parsed_json.get(section, "")
    if isinstance(section_content, (list, dict)):
        section_content = json.dumps(section_content, indent=2)
    elif not isinstance(section_content, str):
        section_content = str(section_content)

    if not section_content.strip():
        return f"No significant content found for the '{section_title}' section in the parsed resume. Please select a section with relevant data to generate questions."

    prompt = f"""Based on the following {section_title} section from the resume: {section_content}
Generate 3 interview questions each for these levels: Generic, Basic, Intermediate, Difficult.
**IMPORTANT: Format the output strictly as follows, with level headers and questions starting with 'Qx:':**
[Generic]
Q1: Question text...
Q2: Question text...
Q3: Question text...
[Basic]
Q1: Question text...
...
[Difficult]
Q3: Question text...
    """
    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.5
    )
    return response.choices[0].message.content.strip()


# --- NEW FUNCTION: JD CHATBOT Q&A ---
def qa_on_jd(question, selected_jd_name):
    """Chatbot for JD (Q&A) using LLM."""
    if not GROQ_API_KEY:
        return "AI Chatbot Disabled: GROQ_API_KEY not set."

    # Find the JD content from the stored list
    jd_item = next((jd for jd in st.session_state.candidate_jd_list if jd['name'] == selected_jd_name), None)

    if not jd_item:
        return "Error: Could not find the selected Job Description in the loaded list."

    jd_text = jd_item['content']
    jd_metadata = {k: v for k, v in jd_item.items() if k not in ['name', 'content']}

    prompt = f"""Given the following Job Description and its extracted metadata:
    
    Job Description Title: {selected_jd_name}
    JD Metadata (JSON): {json.dumps(jd_metadata, indent=2)}
    JD Full Text:
    ---
    {jd_text}
    ---
    
    Answer the following question about the Job Description concisely and directly.
    If the information is not present in the provided text, state that clearly.
    
    Question: {question}
    """
    
    response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.4)
    return response.choices[0].message.content.strip()
# --- END NEW FUNCTION ---


# -------------------------
# Utility Functions
# -------------------------
def dump_to_excel(parsed_json, filename):
    """Dumps parsed JSON data to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profile Data"
    ws.append(["Category", "Details"])
    
    section_order = ['name', 'email', 'phone', 'github', 'linkedin', 'experience', 'education', 'skills', 'projects', 'certifications', 'strength', 'personal_details']
    
    for section_key in section_order:
        if section_key in parsed_json and parsed_json[section_key]:
            content = parsed_json[section_key]
            
            if section_key in ['name', 'email', 'phone', 'github', 'linkedin']:
                ws.append([section_key.replace('_', ' ').title(), str(content)])
            else:
                ws.append([])
                ws.append([section_key.replace('_', ' ').title()])
                
                if isinstance(content, list):
                    for item in content:
                        if item:
                            ws.append(["", str(item)])
                elif isinstance(content, dict):
                    for k, v in content.items():
                        if v:
                            ws.append(["", f"{k.replace('_', ' ').title()}: {v}"])
                    
                else:
                    ws.append(["", str(content)])

    wb.save(filename)
    with open(filename, "rb") as f:
        return f.read()

def parse_and_store_resume(file_input, file_name_key='default', source_type='file'):
    """
    Handles file/text input, parsing, and stores results.
    
    file_input: UploadedFile object or raw text string.
    source_type: 'file' or 'text'.
    """
    
    text = None
    file_name = f"Pasted Text ({date.today().strftime('%Y-%m-%d')})"

    if source_type == 'file':
        if not isinstance(file_input, UploadedFile):
            return {"error": "Invalid file input type passed to parser.", "full_text": ""}

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, file_input.name) 
        with open(temp_path, "wb") as f:
            f.write(file_input.getbuffer()) 

        file_type = get_file_type(temp_path)
        text = extract_content(file_type, temp_path)
        file_name = file_input.name.split('.')[0]
    
    elif source_type == 'text':
        text = file_input
        file_name = f"Pasted Text ({date.today().strftime('%Y-%m-%d')})"
        
    if text.startswith("Error"):
        return {"error": text, "full_text": text, "name": file_name}

    parsed = parse_with_llm(text, return_type='json')
    
    if not parsed or "error" in parsed:
        return {"error": parsed.get('error', 'Unknown parsing error'), "full_text": text, "name": file_name}

    # Generate Excel data for download if needed 
    excel_data = None
    if file_name_key == 'single_resume_candidate':
        try:
            name = parsed.get('name', 'candidate').replace(' ', '_').strip()
            name = "".join(c for c in name if c.isalnum() or c in ('_', '-')).rstrip()
            if not name: name = "candidate"
            excel_filename = os.path.join(tempfile.gettempdir(), f"{name}_parsed_data.xlsx")
            excel_data = dump_to_excel(parsed, excel_filename)
        except Exception as e:
            pass
    
    # Use parsed name if available, otherwise use the generated file_name
    final_name = parsed.get('name', file_name)

    return {
        "parsed": parsed,
        "full_text": text,
        "excel_data": excel_data,
        "name": final_name
    }


def qa_on_resume(question):
    """Chatbot for Resume (Q&A) using LLM."""
    if not GROQ_API_KEY:
        return "AI Chatbot Disabled: GROQ_API_KEY not set."
        
    parsed_json = st.session_state.parsed
    full_text = st.session_state.full_text
    prompt = f"""Given the following resume information:
    Resume Text: {full_text}
    Parsed Resume Data (JSON): {json.dumps(parsed_json, indent=2)}
    Answer the following question about the resume concisely and directly.
    If the information is not present, state that clearly.
    Question: {question}
    """
    response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.4)
    return response.choices[0].message.content.strip()
    # -------------------------
# UI PAGES: Authentication (Login, Signup)
# -------------------------

def login_page():
    st.title("🌐 PragyanAI Job Portal")
    st.header("Login")

    # --- Role Selection ADDED HERE (NEW) ---
    selected_role = st.selectbox(
        "Select Your Role",
        ["Select Role", "Admin Dashboard", "Candidate Dashboard", "Hiring Company Dashboard"],
        key="login_role_select"
    )
    
    st.markdown("---")

    email = st.text_input("Email")
    password = st.text_input("Password", type="password")

    if st.button("Login", use_container_width=True):
        if email and password:
            if selected_role == "Select Role":
                st.error("Please select your role before logging in.")
            elif selected_role == "Admin Dashboard":
                st.success("Login successful! Redirecting to Admin Dashboard.")
                go_to("admin_dashboard")
            elif selected_role == "Candidate Dashboard":
                st.success("Login successful! Redirecting to Candidate Dashboard.")
                go_to("candidate_dashboard")
            elif selected_role == "Hiring Company Dashboard":
                st.success("Login successful! Redirecting to Hiring Company Dashboard.")
                go_to("hiring_dashboard")
        else:
            st.error("Please enter both email and password")

    st.markdown("---")
    
    if st.button("Don't have an account? Sign up here"):
        go_to("signup")

def signup_page():
    st.header("Create an Account")
    email = st.text_input("Email")
    password = st.text_input("Password", type="password")
    confirm = st.text_input("Confirm Password", type="password")

    if st.button("Sign Up", use_container_width=True):
        if password == confirm and email:
            st.success("Signup successful! Please login.")
            go_to("login")
        else:
            st.error("Passwords do not match or email is empty")

    if st.button("Already have an account? Login here"):
        go_to("login")
        # -------------------------
# UI PAGES: Admin and Hiring Dashboards (Kept for context)
# -------------------------

def update_resume_status(user_email, resume_name, new_status, applied_jd, submitted_date):
    """
    Callback function to update the status and metadata of a specific resume,
    now keyed by both user_email and resume_name.
    """
    # 1. Update the overall status dictionary
    # The key is now a combination of email and name for unique identification
    resume_id = (user_email, resume_name)
    st.session_state.resume_statuses[resume_id] = new_status
    
    # 2. Update the resume's metadata within the master list (resumes_to_analyze)
    # We must iterate to find the correct, unique resume entry
    found = False
    for i, resume_data in enumerate(st.session_state.resumes_to_analyze):
        # We assume the Admin Dashboard stores the 'user_email' in the resume data
        if resume_data.get('user_email') == user_email and resume_data['name'] == resume_name:
            st.session_state.resumes_to_analyze[i]['applied_jd'] = applied_jd
            st.session_state.resumes_to_analyze[i]['submitted_date'] = submitted_date
            found = True
            break
            
    if found:
        st.success(f"Status and metadata for **{resume_name}** (User: {user_email}) updated to **{new_status}**.")
    else:
        st.error(f"Error: Could not find resume '{resume_name}' for user '{user_email}' for update.")
        
# --- NEWLY ISOLATED FUNCTIONS FOR APPROVAL TABS ---

def candidate_approval_tab_content():
    st.header("👤 Candidate Approval")
    st.markdown("### Resume Status List (Grouped by Candidate Email)")
    
    if "resumes_to_analyze" not in st.session_state or not st.session_state.resumes_to_analyze:
        st.info("No resumes have been uploaded and parsed in the 'Resume Analysis' tab yet.")
        return
        
    jd_options = [item['name'].replace("--- Simulated JD for: ", "") for item in st.session_state.admin_jd_list]
    jd_options.insert(0, "Select JD") 

    # --- Group resumes by User Email (NEW LOGIC) ---
    resumes_by_user = {}
    for resume_data in st.session_state.resumes_to_analyze:
        # Use a default email if not set, but enforce a real email in a production setup
        user_email = resume_data.get('user_email', 'unknown_user@portal.com') 
        if user_email not in resumes_by_user:
            resumes_by_user[user_email] = []
        resumes_by_user[user_email].append(resume_data)
    # --- End Grouping Logic ---

    
    # Iterate through users and their resumes
    for user_email, resume_list in resumes_by_user.items():
        st.subheader(f"Candidate: **{user_email}**")
        
        for idx, resume_data in enumerate(resume_list):
            resume_name = resume_data['name']
            
            # Use the new tuple key (email, name) for status lookup
            resume_id = (user_email, resume_name)
            current_status = st.session_state.resume_statuses.get(resume_id, "Pending")
            
            current_applied_jd = resume_data.get('applied_jd', 'N/A (Pending Assignment)')
            current_submitted_date = resume_data.get('submitted_date', date.today().strftime("%Y-%m-%d"))

            # Use a unique key suffix based on both email and resume name
            key_suffix = f"{user_email}_{resume_name}_{idx}".replace('.', '_').replace('@', '_')

            with st.container(border=True):
                # 🚨 NEW: Display the Resume Name clearly
                st.markdown(f"**Resume Selection:** **{resume_name}**")
                
                col_jd_input, col_date_input = st.columns(2)
                
                with col_jd_input:
                    try:
                        default_value = current_applied_jd if current_applied_jd != "N/A (Pending Assignment)" else "Select JD"
                        jd_default_index = jd_options.index(default_value)
                    except ValueError:
                        jd_default_index = 0
                        
                    new_applied_jd = st.selectbox(
                        "Applied for JD Title", 
                        options=jd_options,
                        index=jd_default_index,
                        key=f"jd_select_{key_suffix}",
                    )
                    
                with col_date_input:
                    try:
                        date_obj = date.fromisoformat(current_submitted_date)
                    except (ValueError, TypeError):
                        date_obj = date.today()
                        
                    new_submitted_date = st.date_input(
                        "Submitted Date", 
                        value=date_obj,
                        key=f"date_input_{key_suffix}"
                    )
                    
                st.markdown(f"**Current Status:** **{current_status}**")
                
                st.markdown("---")
                
                col1, col2 = st.columns([3, 1])
                
                with col1:
                    st.markdown("Set Status:")
                    new_status = st.selectbox(
                        "Set Status",
                        ["Pending", "Approved", "Rejected", "Shortlisted"],
                        index=["Pending", "Approved", "Rejected", "Shortlisted"].index(current_status),
                        key=f"status_select_{key_suffix}",
                        label_visibility="collapsed"
                    )

                with col2:
                    if st.button("Update", key=f"update_btn_{key_suffix}"):
                        
                        if new_applied_jd == "Select JD" and len(jd_options) > 1:
                            jd_to_save = "N/A (Pending Assignment)"
                        else:
                            jd_to_save = new_applied_jd
                            
                        # 🚨 CRITICAL: Call the updated function with user_email
                        update_resume_status(
                            user_email,
                            resume_name, 
                            new_status, 
                            jd_to_save, 
                            new_submitted_date.strftime("%Y-%m-%d")
                        )
                        st.rerun() 
            
    st.markdown("---")
            
    summary_data = []
    # Loop over the raw data to generate summary (easier than reconstructing)
    for resume_data in st.session_state.resumes_to_analyze:
        name = resume_data['name']
        user_email = resume_data.get('user_email', 'unknown_user@portal.com')
        resume_id = (user_email, name)
        
        summary_data.append({
            "User Email": user_email, # 🚨 ADDED USER EMAIL
            "Resume": name, 
            "Applied JD": resume_data.get('applied_jd', 'N/A'),
            "Submitted Date": resume_data.get('submitted_date', 'N/A'),
            "Status": st.session_state.resume_statuses.get(resume_id, "Pending")
        })
        
    st.subheader("Summary of All Resumes")
    st.dataframe(summary_data, use_container_width=True)


def vendor_approval_tab_content():
    # ... (No changes to this function) ...
    st.header("🤝 Vendor Approval") 
    
    st.markdown("### 1. Add New Vendor")
    if "vendors" not in st.session_state:
        st.session_state.vendors = []
    if "vendor_statuses" not in st.session_state:
        st.session_state.vendor_statuses = {}
        
    with st.form("add_vendor_form"):
        col1, col2 = st.columns(2)
        with col1:
            vendor_name = st.text_input("Vendor Name", key="new_vendor_name")
        with col2:
            vendor_domain = st.text_input("Service / Domain Name", key="new_vendor_domain")
            
        col3, col4 = st.columns(2)
        with col3:
            submitted_date = st.date_input("Submitted Date", value=date.today(), key="new_vendor_date")
        with col4:
            initial_status = st.selectbox(
                "Set Status", 
                ["Pending Review", "Approved", "Rejected"],
                key="new_vendor_status"
            )
        
        add_vendor_button = st.form_submit_button("Add Vendor", use_container_width=True)

        if add_vendor_button:
            if vendor_name and vendor_domain:
                vendor_id = vendor_name.strip() 
                
                if vendor_id in st.session_state.vendor_statuses:
                    st.warning(f"Vendor '{vendor_name}' already exists.")
                else:
                    new_vendor = {
                        'name': vendor_name.strip(),
                        'domain': vendor_domain.strip(),
                        'submitted_date': submitted_date.strftime("%Y-%m-%d")
                    }
                    st.session_state.vendors.append(new_vendor)
                    st.session_state.vendor_statuses[vendor_id] = initial_status
                    st.success(f"Vendor **{vendor_name}** added successfully with status **{initial_status}**.")
                    st.rerun() 
            else:
                st.error("Please fill in both Vendor Name and Service / Domain Name.")

    st.markdown("---")
    
    st.markdown("### 2. Update Existing Vendor Status")
    
    if not st.session_state.vendors:
        st.info("No vendors have been added yet.")
    else:
        for idx, vendor in enumerate(st.session_state.vendors):
            vendor_name = vendor['name']
            vendor_id = vendor_name 
            current_status = st.session_state.vendor_statuses.get(vendor_id, "Unknown")
            
            with st.container(border=True):
                
                col_info, col_status_input, col_update_btn = st.columns([3, 2, 1])
                
                with col_info:
                    st.markdown(f"**Vendor:** {vendor_name} (`{vendor['domain']}`) - *Submitted: {vendor['submitted_date']}*")
                    st.markdown(f"**Current Status:** **{current_status}**")
                    
                with col_status_input:
                    new_status = st.selectbox(
                        "Set Status",
                        ["Pending Review", "Approved", "Rejected"],
                        index=["Pending Review", "Approved", "Rejected"].index(current_status),
                        key=f"vendor_status_select_{idx}",
                        label_visibility="collapsed"
                    )

                with col_update_btn:
                    st.markdown("##") 
                    if st.button("Update", key=f"vendor_update_btn_{idx}", use_container_width=True):
                        
                        st.session_state.vendor_statuses[vendor_id] = new_status
                        
                        st.success(f"Status for **{vendor_name}** updated to **{new_status}**.")
                        st.rerun()
                        
        st.markdown("---")
        
        summary_data = []
        for vendor in st.session_state.vendors:
            name = vendor['name']
            summary_data.append({
                "Vendor Name": name,
                "Domain": vendor['domain'],
                "Submitted Date": vendor['submitted_date'],
                "Status": st.session_state.vendor_statuses.get(name, "Unknown")
            })
        
        st.subheader("Summary of All Vendors")
        st.dataframe(summary_data, use_container_width=True)

def admin_dashboard():
    st.header("🧑‍💼 Admin Dashboard")
    
    # --- NAVIGATION BLOCK (MODIFIED) ---
    nav_col, _ = st.columns([1, 1]) 

    with nav_col:
        if st.button("🚪 Log Out", use_container_width=True):
            go_to("login") 
    # --- END NAVIGATION BLOCK ---
    
    # Initialize Admin session state variables (Defensive check)
    if "admin_jd_list" not in st.session_state: st.session_state.admin_jd_list = []
    if "resumes_to_analyze" not in st.session_state: st.session_state.resumes_to_analyze = []
    if "admin_match_results" not in st.session_state: st.session_state.admin_match_results = []
    # 🚨 UPDATED: resume_statuses key is now a tuple (user_email, resume_name)
    if "resume_statuses" not in st.session_state: st.session_state.resume_statuses = {} 
    if "vendors" not in st.session_state: st.session_state.vendors = []
    if "vendor_statuses" not in st.session_state: st.session_state.vendor_statuses = {}
        
    
    # --- TAB ORDER ---
    tab_jd, tab_analysis, tab_user_mgmt, tab_statistics = st.tabs([
        "📄 JD Management", 
        "📊 Resume Analysis", 
        "🛠️ User Management", 
        "📈 Statistics" 
    ])
    # -------------------------

    # --- TAB 1: JD Management ---
    with tab_jd:
        st.subheader("Add and Manage Job Descriptions (JD)")
        
        jd_type = st.radio("Select JD Type", ["Single JD", "Multiple JD"], key="jd_type_admin")
        st.markdown("### Add JD by:")
        
        method = st.radio("Choose Method", ["Upload File", "Paste Text", "LinkedIn URL"], key="jd_add_method_admin") 

        # URL
        if method == "LinkedIn URL":
            url_list = st.text_area(
                "Enter one or more URLs (comma separated)" if jd_type == "Multiple JD" else "Enter URL", key="url_list_admin"
            )
            if st.button("Add JD(s) from URL", key="add_jd_url_btn_admin"):
                if url_list:
                    urls = [u.strip() for u in url_list.split(",")] if jd_type == "Multiple JD" else [url_list.strip()]
                    
                    count = 0
                    for url in urls:
                        if not url: continue
                        
                        with st.spinner(f"Attempting JD extraction for: {url}"):
                            jd_text = extract_jd_from_linkedin_url(url)
                            metadata = extract_jd_metadata(jd_text) # NEW METADATA EXTRACTION
                        
                        name_base = url.split('/jobs/view/')[-1].split('/')[0] if '/jobs/view/' in url else f"URL {count+1}"
                        st.session_state.admin_jd_list.append({"name": f"JD from URL: {name_base}", "content": jd_text, **metadata}) # ADD METADATA
                        if not jd_text.startswith("[Error"):
                            count += 1
                            
                    if count > 0:
                        st.success(f"✅ {count} JD(s) added successfully! Check the display below for the extracted content.")
                    else:
                        st.error("No JDs were added successfully.")


        # Paste Text
        elif method == "Paste Text":
            text_list = st.text_area(
                "Paste one or more JD texts (separate by '---')" if jd_type == "Multiple JD" else "Paste JD text here", key="text_list_admin"
            )
            if st.button("Add JD(s) from Text", key="add_jd_text_btn_admin"):
                if text_list:
                    texts = [t.strip() for t in text_list.split("---")] if jd_type == "Multiple JD" else [text_list.strip()]
                    for i, text in enumerate(texts):
                         if text:
                            name_base = text.splitlines()[0].strip()
                            if len(name_base) > 30: name_base = f"{name_base[:27]}..."
                            if not name_base: name_base = f"Pasted JD {len(st.session_state.admin_jd_list) + i + 1}"
                            
                            metadata = extract_jd_metadata(text) # NEW METADATA EXTRACTION
                            st.session_state.admin_jd_list.append({"name": name_base, "content": text, **metadata}) # ADD METADATA
                    st.success(f"✅ {len(texts)} JD(s) added successfully!")

        # Upload File
        elif method == "Upload File":
            uploaded_files = st.file_uploader(
                "Upload JD file(s)",
                type=["pdf", "txt", "docx"],
                accept_multiple_files=(jd_type == "Multiple JD"), # Dynamically set
                key="jd_file_uploader_admin"
            )
            
            if st.button("Add JD(s) from File", key="add_jd_file_btn_admin"):
                # CRITICAL FIX: Ensure 'files_to_process' is always a list of single UploadedFile objects
                if uploaded_files is None:
                    st.warning("Please upload file(s).")
                    
                files_to_process = uploaded_files if isinstance(uploaded_files, list) else ([uploaded_files] if uploaded_files else [])
                
                count = 0
                for file in files_to_process:
                    if file: # Check if file object is not None
                        temp_dir = tempfile.mkdtemp()
                        temp_path = os.path.join(temp_dir, file.name)
                        with open(temp_path, "wb") as f:
                            f.write(file.getbuffer())
                            
                        file_type = get_file_type(temp_path)
                        jd_text = extract_content(file_type, temp_path)
                        
                        if not jd_text.startswith("Error"):
                            metadata = extract_jd_metadata(jd_text) # NEW METADATA EXTRACTION
                            st.session_state.admin_jd_list.append({"name": file.name, "content": jd_text, **metadata}) # ADD METADATA
                            count += 1
                        else:
                            st.error(f"Error extracting content from {file.name}: {jd_text}")
                            
                if count > 0:
                    st.success(f"✅ {count} JD(s) added successfully!")
                elif uploaded_files:
                    st.error("No valid JD files were uploaded or content extraction failed.")


        # Display Added JDs
        if st.session_state.admin_jd_list:
            
            col_display_header, col_clear_button = st.columns([3, 1])
            
            with col_display_header:
                st.markdown("### ✅ Current JDs Added:")
                
            with col_clear_button:
                if st.button("🗑️ Clear All JDs", key="clear_jds_admin", use_container_width=True, help="Removes all currently loaded JDs."):
                    st.session_state.admin_jd_list = []
                    st.session_state.admin_match_results = [] 
                    st.success("All JDs and associated match results have been cleared.")
                    st.rerun() 

            for idx, jd_item in enumerate(st.session_state.admin_jd_list, 1):
                title = jd_item['name']
                display_title = title.replace("--- Simulated JD for: ", "")
                with st.expander(f"JD {idx}: {display_title} | Role: {jd_item.get('role', 'N/A')}"):
                    st.markdown(f"**Job Type:** {jd_item.get('job_type', 'N/A')} | **Key Skills:** {', '.join(jd_item.get('key_skills', ['N/A']))}") # ADDED METADATA DISPLAY
                    st.markdown("---")
                    st.text(jd_item['content'])
        else:
            st.info("No Job Descriptions added yet.")


    # --- TAB 2: Resume Analysis --- 
    with tab_analysis:
        st.subheader("Analyze Resumes Against Job Descriptions")

        # 1. Resume Upload
        st.markdown("#### 1. Upload Resumes")
        
        # 🚨 NEW: Add Email Input for the Uploader
        uploaded_email = st.text_input(
            "Enter Candidate Email for uploaded resumes (Mandatory for multi-user system)",
            value=st.session_state.get('last_uploaded_email_admin', 'test_user@portal.com'),
            key="admin_resume_upload_email"
        )
        st.session_state.last_uploaded_email_admin = uploaded_email
        
        resume_upload_type = st.radio("Upload Type", ["Single Resume", "Multiple Resumes"], key="resume_upload_type_admin")

        uploaded_files = st.file_uploader(
            "Choose files to analyze",
            type=["pdf", "docx", "txt", "json", "rtf"], 
            accept_multiple_files=(resume_upload_type == "Multiple Resumes"),
            key="resume_file_uploader_admin"
        )
        
        col_parse, col_clear = st.columns([3, 1])
        
        with col_parse:
            if st.button("Load and Parse Resume(s) for Analysis", key="parse_resumes_admin", use_container_width=True):
                if not uploaded_email.strip():
                    st.error("Please enter a **Candidate Email** before parsing resumes.")
                    return
                
                if uploaded_files:
                    # CRITICAL FIX: Ensure 'files_to_process' is always a list of single UploadedFile objects
                    files_to_process = uploaded_files if isinstance(uploaded_files, list) else ([uploaded_files] if uploaded_files else [])
                    
                    count = 0
                    with st.spinner("Parsing resume(s)... This may take a moment."):
                        for file in files_to_process:
                            if file: # Check if file object is not None
                                # The loop ensures 'file' is a single UploadedFile object.
                                result = parse_and_store_resume(file, file_name_key='admin_analysis', source_type='file')
                                
                                if "error" not in result:
                                    # 🚨 CRITICAL: Store the user_email with the resume data
                                    result['user_email'] = uploaded_email 
                                    result['applied_jd'] = "N/A (Pending Assignment)"
                                    result['submitted_date'] = date.today().strftime("%Y-%m-%d")
                                    
                                    st.session_state.resumes_to_analyze.append(result)
                                    
                                    # 🚨 UPDATED: Status key is now (email, name)
                                    resume_id = (uploaded_email, result['name'])
                                    if resume_id not in st.session_state.resume_statuses:
                                        st.session_state.resume_statuses[resume_id] = "Pending"
                                    
                                    count += 1
                                else:
                                    st.error(f"Failed to parse {file.name}: {result['error']}")

                    if count > 0:
                        st.success(f"Successfully loaded and parsed {count} resume(s) for analysis under email: **{uploaded_email}**.")
                        st.rerun() 
                    elif not st.session_state.resumes_to_analyze:
                        st.warning("No resumes were successfully loaded and parsed.")
                else:
                    st.warning("Please upload one or more resume files.")
        
        with col_clear:
            if st.button("🗑️ Clear All Resumes", key="clear_resumes_admin", use_container_width=True, help="Removes all currently loaded resumes and match results."):
                st.session_state.resumes_to_analyze = []
                st.session_state.admin_match_results = []
                st.session_state.resume_statuses = {} 
                st.success("All resumes and associated match results have been cleared.")
                st.rerun() 


        st.markdown("---")

        # 2. JD Selection and Analysis
        st.markdown("#### 2. Select JD and Run Analysis")

        if not st.session_state.resumes_to_analyze:
            st.info("Upload and parse resumes first to enable analysis.")
            if not st.session_state.admin_jd_list: return
        elif not st.session_state.admin_jd_list:
            st.error("Please add at least one Job Description in the 'JD Management' tab before running an analysis.")
            return

        # 🚨 UPDATED: Display resumes with their associated email
        resume_options = [
            f"{r['name']} (User: {r.get('user_email', 'unknown')})" 
            for r in st.session_state.resumes_to_analyze
        ]
        
        selected_resume_options = st.multiselect(
            "Select Resume(s) for Matching",
            options=resume_options,
            default=resume_options, 
            key="select_resumes_admin"
        )
        
        # 🚨 NEW: Map selected options back to the actual data structure
        resumes_to_match = []
        for option in selected_resume_options:
            name_part = option.split(' (User: ')[0]
            email_part = option.split(' (User: ')[1].rstrip(')')
            
            # Find the actual resume data entry
            found_resume = next((
                r for r in st.session_state.resumes_to_analyze 
                if r['name'] == name_part and r.get('user_email', 'unknown') == email_part
            ), None)
            
            if found_resume:
                resumes_to_match.append(found_resume)

        jd_options = {item['name']: item['content'] for item in st.session_state.admin_jd_list}
        selected_jd_name = st.selectbox("Select JD for Matching", list(jd_options.keys()), key="select_jd_admin")
        selected_jd_content = jd_options.get(selected_jd_name, "")


        if st.button(f"Run Match Analysis on {len(resumes_to_match)} Selected Resume(s)", key="run_match_analysis_admin"):
            st.session_state.admin_match_results = []
            
            if not selected_jd_content:
                st.error("Selected JD content is empty.")
                return
            
            if not resumes_to_match:
                st.warning("No resumes were selected for matching.")
                return

            results_with_score = []
            
            with st.spinner(f"Matching {len(resumes_to_match)} resumes against '{selected_jd_name}'..."):
                for resume_data in resumes_to_match: 
                    
                    resume_name = resume_data['name']
                    user_email = resume_data.get('user_email', 'unknown_user@portal.com') # Get the email
                    parsed_json = resume_data['parsed']

                    try:
                        fit_output = evaluate_jd_fit(selected_jd_content, parsed_json)
                        
                        overall_score_match = re.search(r'Overall Fit Score:\s*[^\d]*(\d+)\s*/10', fit_output, re.IGNORECASE)
                        section_analysis_match = re.search(
                             r'--- Section Match Analysis ---\s*(.*?)\s*Strengths/Matches:', 
                             fit_output, re.DOTALL
                        )

                        skills_percent, experience_percent, education_percent = 'N/A', 'N/A', 'N/A'
                        
                        if section_analysis_match:
                            section_text = section_analysis_match.group(1)
                            skills_match = re.search(r'Skills Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                            experience_match = re.search(r'Experience Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                            education_match = re.search(r'Education Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                            
                            if skills_match: skills_percent = skills_match.group(1)
                            if experience_match: experience_percent = experience_match.group(1)
                            if education_match: education_percent = education_match.group(1)
                        
                        overall_score = overall_score_match.group(1) if overall_score_match else 'N/A'
                        
                        results_with_score.append({
                            "user_email": user_email, # 🚨 ADDED USER EMAIL
                            "resume_name": resume_name,
                            "jd_name": selected_jd_name,
                            "overall_score": overall_score,
                            "numeric_score": int(overall_score) if overall_score.isdigit() else -1, # For sorting
                            "skills_percent": skills_percent,
                            "experience_percent": experience_percent, 
                            "education_percent": education_percent,   
                            "full_analysis": fit_output
                        })
                    except Exception as e:
                        results_with_score.append({
                            "user_email": user_email, # 🚨 ADDED USER EMAIL
                            "resume_name": resume_name,
                            "jd_name": selected_jd_name,
                            "overall_score": "Error",
                            "numeric_score": -1, # Set a low score for errors
                            "skills_percent": "Error",
                            "experience_percent": "Error", 
                            "education_percent": "Error",   
                            "full_analysis": f"Error running analysis: {e}\n{traceback.format_exc()}"
                        })
                
                # Sort and Rank (Admin Tab doesn't use the Rank field, but this is good practice)
                results_with_score.sort(key=lambda x: x['numeric_score'], reverse=True)
                
                # Store the sorted results
                st.session_state.admin_match_results = results_with_score

            st.success("Analysis complete!")


        # 3. Display Results
        if st.session_state.get('admin_match_results'):
            st.markdown("#### 3. Match Results")
            results_df = st.session_state.admin_match_results
            
            display_data = []
            for item in results_df:
                # 🚨 UPDATED: Status lookup uses tuple (email, name)
                resume_id = (item.get('user_email', 'unknown_user@portal.com'), item["resume_name"])
                status = st.session_state.resume_statuses.get(resume_id, 'Pending') 
                
                display_data.append({
                    "User Email": item["user_email"], # 🚨 ADDED USER EMAIL
                    "Resume": item["resume_name"],
                    "JD": item["jd_name"],
                    "Fit Score (out of 10)": item["overall_score"],
                    "Skills (%)": item.get("skills_percent", "N/A"),
                    "Experience (%)": item.get("experience_percent", "N/A"), 
                    "Education (%)": item.get("education_percent", "N/A"),
                    "Approval Status": status
                })

            st.dataframe(display_data, use_container_width=True)

            st.markdown("##### Detailed Reports")
            for item in results_df:
                # 🚨 UPDATED: Header includes email
                header_text = f"Report for **{item['resume_name']}** (User: {item['user_email']}) against {item['jd_name']} (Score: **{item['overall_score']}/10** | S: **{item.get('skills_percent', 'N/A')}%** | E: **{item.get('experience_percent', 'N/A')}%** | Edu: **{item.get('education_percent', 'N/A')}%**)"
                with st.expander(header_text):
                    st.markdown(item['full_analysis'])

                    
    # --- TAB 3: User Management (NEW PARENT TAB) ---
    with tab_user_mgmt:
        st.header("🛠️ User Management")
        
        nested_tab_candidate, nested_tab_vendor = st.tabs([
            "👤 Candidate Approval",
            "🤝 Vendor Approval"
        ])
        
        with nested_tab_candidate:
            candidate_approval_tab_content() 
            
        with nested_tab_vendor:
            vendor_approval_tab_content() 


    # --- TAB 4: Statistics (Renumbered) ---
    with tab_statistics:
        st.header("System Statistics")
        st.markdown("---")

        total_candidates = len(st.session_state.resumes_to_analyze)
        total_jds = len(st.session_state.admin_jd_list)
        total_vendors = len(st.session_state.vendors)
        no_of_applications = total_candidates 
        
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(label="Total Candidates", value=total_candidates, delta="Resumes Submitted")

        with col2:
            st.metric(label="Total JDs", value=total_jds, delta_color="off")
        
        with col3:
            st.metric(label="Total Vendors", value=total_vendors, delta_color="off")

        with col4:
            st.metric(label="No. of Applications", value=no_of_applications, delta_color="off")
            
        st.markdown("---")
        
        # 🚨 UPDATED: Loop over tuple keys to get statuses
        status_counts = {}
        for status in st.session_state.resume_statuses.values():
            status_counts[status] = status_counts.get(status, 0) + 1
            
        st.subheader("Candidate Status Breakdown")
        
        status_cols = st.columns(len(status_counts) or 1)
        
        if status_counts:
            col_count = len(status_cols)
            for i, (status, count) in enumerate(status_counts.items()):
                with status_cols[i % col_count]:
                    st.metric(label=f"{status}", value=count)
        else:
            st.info("No resumes loaded to calculate status breakdown.")
            # --- NEW HELPER FUNCTION FOR HTML/PDF Generation ---
def generate_cv_html(parsed_data):
    """Generates a simple, print-friendly HTML string from parsed data for PDF conversion."""
    
    # Simple CSS for a clean, print-friendly CV look
    css = """
    <style>
        @page { size: A4; margin: 1cm; }
        body { font-family: 'Arial', sans-serif; line-height: 1.5; margin: 0; padding: 0; font-size: 10pt; }
        .header { text-align: center; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 20px; }
        .header h1 { margin: 0; font-size: 1.8em; }
        .contact-info { display: flex; justify-content: center; font-size: 0.8em; color: #555; }
        .contact-info span { margin: 0 8px; }
        .section { margin-bottom: 15px; page-break-inside: avoid; }
        .section h2 { border-bottom: 1px solid #999; padding-bottom: 3px; margin-bottom: 8px; font-size: 1.1em; text-transform: uppercase; color: #333; }
        .item-list ul { list-style-type: disc; margin-left: 20px; padding-left: 0; margin-top: 0; }
        .item-list ul li { margin-bottom: 3px; }
        .item-list p { margin: 3px 0 8px 0; }
        a { color: #0056b3; text-decoration: none; }
    </style>
    """
    
    # --- HTML Structure ---
    html_content = f"<html><head>{css}<title>{parsed_data.get('name', 'CV')}</title></head><body>"
    
    # 1. Header and Contact Info
    html_content += '<div class="header">'
    html_content += f"<h1>{parsed_data.get('name', 'Candidate Name')}</h1>"
    
    contact_parts = []
    if parsed_data.get('email'): contact_parts.append(f"<span>📧 {parsed_data['email']}</span>")
    if parsed_data.get('phone'): contact_parts.append(f"<span>📱 {parsed_data['phone']}</span>")
    if parsed_data.get('linkedin'): contact_parts.append(f"<span>🔗 <a href='{parsed_data['linkedin']}'>{parsed_data.get('linkedin', 'LinkedIn').split('/')[-1] if parsed_data.get('linkedin') else 'LinkedIn'}</a></span>")
    if parsed_data.get('github'): contact_parts.append(f"<span>💻 <a href='{parsed_data['github']}'>{parsed_data.get('github', 'GitHub').split('/')[-1] if parsed_data.get('github') else 'GitHub'}</a></span>")
    
    html_content += f'<div class="contact-info">{" | ".join(contact_parts)}</div>'
    html_content += '</div>'
    
    # 2. Sections
    section_order = ['personal_details', 'experience', 'projects', 'education', 'certifications', 'skills', 'strength']
    
    for k in section_order:
        v = parsed_data.get(k)
        
        # Skip contact details already handled
        if k in ['name', 'email', 'phone', 'linkedin', 'github']: continue 

        if v and (isinstance(v, str) and v.strip() or isinstance(v, list) and v):
            
            html_content += f'<div class="section"><h2>{k.replace("_", " ").title()}</h2>'
            html_content += '<div class="item-list">'
            
            if k == 'personal_details' and isinstance(v, str):
                html_content += f"<p>{v}</p>"
            elif isinstance(v, list):
                html_content += '<ul>'
                for item in v:
                    if item: 
                        html_content += f"<li>{item}</li>"
                html_content += '</ul>'
            else:
                html_content += f"<p>{v}</p>"
                
            html_content += '</div></div>'

    html_content += '</body></html>'
    return html_content


def cv_management_tab_content():
    st.header("📝 Prepare Your CV")
    st.markdown("### 1. Form Based CV Builder")
    st.info("Fill out the details below to generate a parsed CV that can be used immediately for matching and interview prep, or start by parsing a file in the 'Resume Parsing' tab.")

    # Initialize the parsed data if not already existing
    default_parsed = {
        "name": "", "email": "", "phone": "", "linkedin": "", "github": "",
        "skills": [], "experience": [], "education": [], "certifications": [], 
        "projects": [], "strength": [], "personal_details": ""
    }
    
    # Use a specific session state key for form data, initializing from parsed if available
    if "cv_form_data" not in st.session_state:
        # Load existing parsed data or default if the tab is opened for the first time
        if st.session_state.get('parsed', {}).get('name'):
            st.session_state.cv_form_data = st.session_state.parsed.copy()
        else:
            st.session_state.cv_form_data = default_parsed
    
    # --- CV Builder Form ---
    with st.form("cv_builder_form"):
        st.subheader("Personal & Contact Details")
        
        # Row 1: Name, Email, Phone
        col1, col2, col3 = st.columns(3)
        with col1:
            st.session_state.cv_form_data['name'] = st.text_input(
                "Full Name", 
                value=st.session_state.cv_form_data['name'], 
                key="cv_name"
            )
        with col2:
            st.session_state.cv_form_data['email'] = st.text_input(
                "Email Address", 
                value=st.session_state.cv_form_data['email'], 
                key="cv_email"
            )
        with col3:
            st.session_state.cv_form_data['phone'] = st.text_input(
                "Phone Number", 
                value=st.session_state.cv_form_data['phone'], 
                key="cv_phone"
            )
        
        # Row 2: LinkedIn, GitHub
        col4, col5 = st.columns(2)
        with col4:
            st.session_state.cv_form_data['linkedin'] = st.text_input(
                "LinkedIn Profile URL", 
                value=st.session_state.cv_form_data.get('linkedin', ''), 
                key="cv_linkedin"
            )
        with col5:
            st.session_state.cv_form_data['github'] = st.text_input(
                "GitHub Profile URL", 
                value=st.session_state.cv_form_data.get('github', ''), 
                key="cv_github"
            )
        
        # Row 3: Summary/Personal Details 
        st.markdown("---")
        st.subheader("Summary / Personal Details")
        st.session_state.cv_form_data['personal_details'] = st.text_area(
            "Professional Summary or Personal Details (e.g., date of birth, address, nationality)", 
            value=st.session_state.cv_form_data.get('personal_details', ''), 
            height=100,
            key="cv_personal_details"
        )
        
        st.markdown("---")
        st.subheader("Technical Sections (One Item per Line)")

        # Skills
        skills_text = "\n".join(st.session_state.cv_form_data.get('skills', []))
        new_skills_text = st.text_area(
            "Key Skills (Technical and Soft)", 
            value=skills_text,
            height=150,
            key="cv_skills"
        )
        st.session_state.cv_form_data['skills'] = [s.strip() for s in new_skills_text.split('\n') if s.strip()]
        
        # Experience
        experience_text = "\n".join(st.session_state.cv_form_data.get('experience', []))
        new_experience_text = st.text_area(
            "Professional Experience (Job Roles, Companies, Dates, Key Responsibilities)", 
            value=experience_text,
            height=150,
            key="cv_experience"
        )
        st.session_state.cv_form_data['experience'] = [e.strip() for e in new_experience_text.split('\n') if e.strip()]

        # Education
        education_text = "\n".join(st.session_state.cv_form_data.get('education', []))
        new_education_text = st.text_area(
            "Education (Degrees, Institutions, Dates)", 
            value=education_text,
            height=100,
            key="cv_education"
        )
        st.session_state.cv_form_data['education'] = [d.strip() for d in new_education_text.split('\n') if d.strip()]
        
        # Certifications
        certifications_text = "\n".join(st.session_state.cv_form_data.get('certifications', []))
        new_certifications_text = st.text_area(
            "Certifications (Name, Issuing Body, Date)", 
            value=certifications_text,
            height=100,
            key="cv_certifications"
        )
        st.session_state.cv_form_data['certifications'] = [c.strip() for c in new_certifications_text.split('\n') if c.strip()]
        
        # Projects
        projects_text = "\n".join(st.session_state.cv_form_data.get('projects', []))
        new_projects_text = st.text_area(
            "Projects (Name, Description, Technologies)", 
            value=projects_text,
            height=150,
            key="cv_projects"
        )
        st.session_state.cv_form_data['projects'] = [p.strip() for p in new_projects_text.split('\n') if p.strip()]
        
        # Strengths
        strength_text = "\n".join(st.session_state.cv_form_data.get('strength', []))
        new_strength_text = st.text_area(
            "Strengths / Key Personal Qualities (One per line)", 
            value=strength_text,
            height=100,
            key="cv_strength"
        )
        st.session_state.cv_form_data['strength'] = [s.strip() for s in new_strength_text.split('\n') if s.strip()]


        submit_form_button = st.form_submit_button("Generate and Load CV Data", use_container_width=True)

    if submit_form_button:
        # 1. Basic validation
        if not st.session_state.cv_form_data['name'] or not st.session_state.cv_form_data['email']:
            st.error("Please fill in at least your **Full Name** and **Email Address**.")
            return

        # 2. Update the main session state variables (as if a file was parsed)
        st.session_state.parsed = st.session_state.cv_form_data.copy()
        st.session_state.parsed['name'] = st.session_state.cv_form_data['name'] # Ensure name is set for display
        
        # 3. Create a placeholder full_text for Q&A (simple compilation of all fields)
        compiled_text = ""
        for k, v in st.session_state.cv_form_data.items():
            if v:
                compiled_text += f"{k.replace('_', ' ').title()}:\n"
                if isinstance(v, list):
                    compiled_text += "\n".join([f"- {item}" for item in v]) + "\n\n"
                else:
                    compiled_text += str(v) + "\n\n"
        st.session_state.full_text = compiled_text
        
        # 4. Clear related states (since this is a new resume)
        st.session_state.candidate_match_results = []
        st.session_state.interview_qa = []
        st.session_state.evaluation_report = ""

        st.success(f"✅ CV data for **{st.session_state.parsed['name']}** successfully generated and loaded! You can now use the Chatbot, Match, and Interview Prep tabs.")
        
    st.markdown("---")
    st.subheader("2. Loaded CV Data Preview and Download")
    
    # --- TABBED VIEW SECTION (PDF/MARKDOWN/JSON) ---
    if st.session_state.get('parsed', {}).get('name'):
        
        # Filter for non-empty/non-list fields before sending to formatter
        filled_data_for_preview = {
            k: v for k, v in st.session_state.parsed.items() 
            if v and (isinstance(v, str) and v.strip() or isinstance(v, list) and v)
        }
        
        # Helper function for Markdown formatting
        def format_parsed_json_to_markdown(parsed_data):
            """Formats the parsed JSON data into a clean, CV-like Markdown structure."""
            md = ""
            
            # --- Personal Info (Header) ---
            if parsed_data.get('name'):
                md += f"# **{parsed_data['name']}**\n\n"
            
            contact_info = []
            if parsed_data.get('email'): contact_info.append(parsed_data['email'])
            if parsed_data.get('phone'): contact_info.append(parsed_data['phone'])
            if parsed_data.get('linkedin'): contact_info.append(f"[LinkedIn]({parsed_data['linkedin']})")
            if parsed_data.get('github'): contact_info.append(f"[GitHub]({parsed_data['github']})")
            
            if contact_info:
                md += f"| {' | '.join(contact_info)} |\n"
                md += "| " + " | ".join(["---"] * len(contact_info)) + " |\n\n"
            
            # --- Section Content ---
            section_order = ['personal_details', 'experience', 'projects', 'education', 'certifications', 'skills', 'strength']
            
            for k in section_order:
                v = parsed_data.get(k)
                
                # Skip contact details already handled in header
                if k in ['name', 'email', 'phone', 'linkedin', 'github']: continue 

                if v and (isinstance(v, str) and v.strip() or isinstance(v, list) and v):
                    
                    md += f"## **{k.replace('_', ' ').upper()}**\n"
                    md += "---\n"
                    
                    if k == 'personal_details' and isinstance(v, str):
                        md += f"{v}\n\n"
                    elif isinstance(v, list):
                        for item in v:
                            if item: 
                                # Use bullet points for list items (Experience, Skills, Projects, etc.)
                                md += f"- {item}\n"
                        md += "\n"
                    else:
                        # Fallback for any other string
                        md += f"{v}\n\n"
            return md


        tab_markdown, tab_json, tab_pdf = st.tabs(["📝 Markdown View", "💾 JSON View", "⬇️ PDF/HTML Download"])

        # --- Markdown View ---
        with tab_markdown:
            cv_markdown_preview = format_parsed_json_to_markdown(filled_data_for_preview)
            st.markdown(cv_markdown_preview)

            # Markdown Download Button
            st.download_button(
                label="⬇️ Download CV as Markdown (.md)",
                data=cv_markdown_preview,
                file_name=f"{st.session_state.parsed.get('name', 'Generated_CV').replace(' ', '_')}_CV_Document.md",
                mime="text/markdown",
                key="download_cv_markdown_final"
            )


        # --- JSON View ---
        with tab_json:
            st.json(st.session_state.parsed)
            st.info("This is the raw, structured data used by the AI tools.")

            # JSON Download Button
            json_output = json.dumps(st.session_state.parsed, indent=2)
            st.download_button(
                label="⬇️ Download CV as JSON File",
                data=json_output,
                file_name=f"{st.session_state.parsed.get('name', 'Generated_CV').replace(' ', '_')}_CV_Data.json",
                mime="application/json",
                key="download_cv_json_final"
            )


        # --- PDF View (Download) ---
        with tab_pdf:
            st.markdown("### Download CV as HTML (Print-to-PDF)")
            st.info("Click the button below to download an HTML file. Open the file in your browser and use the browser's **'Print'** function, selecting **'Save as PDF'** to create your final CV document.")
            
            html_output = generate_cv_html(filled_data_for_preview)

            st.download_button(
                label="⬇️ Download CV as Print-Ready HTML File (for PDF conversion)",
                data=html_output,
                file_name=f"{st.session_state.parsed.get('name', 'Generated_CV').replace(' ', '_')}_CV_Document.html",
                mime="text/html",
                key="download_cv_html"
            )
            
            st.markdown("---")
            st.markdown("### Raw Text Data Download (for utility)")
            st.download_button(
                label="⬇️ Download All CV Data as Raw Text (.txt)",
                data=st.session_state.full_text,
                file_name=f"{st.session_state.parsed.get('name', 'Generated_CV').replace(' ', '_')}_Raw_Data.txt",
                mime="text/plain",
                key="download_cv_txt_final"
            )
            
    else:
        st.info("Please fill out the form above and click 'Generate and Load CV Data' or parse a resume in the 'Resume Parsing' tab to see the preview and download options.")


def filter_jd_tab_content():
    st.header("🔍 Filter Job Descriptions by Criteria")
    st.markdown("Use the filters below to narrow down your saved Job Descriptions.")

    if not st.session_state.candidate_jd_list:
        st.info("No Job Descriptions are currently loaded. Please add JDs in the 'JD Management' tab (Tab 4).")
        # Ensure filtered list is cleared/initialized
        if 'filtered_jds_display' not in st.session_state:
            st.session_state.filtered_jds_display = []
        return
    
    # --- Skill and Role Extraction (outside the form so options are available immediately) ---
    unique_roles = sorted(list(set(
        [item.get('role', 'General Analyst') for item in st.session_state.candidate_jd_list] + DEFAULT_ROLES
    )))
    unique_job_types = sorted(list(set(
        [item.get('job_type', 'Full-time') for item in st.session_state.candidate_jd_list] + DEFAULT_JOB_TYPES
    )))
    
    STARTER_KEYWORDS = {
        "Python", "MySQL", "GCP", "cloud computing", "ML", 
        "API services", "LLM integration", "JavaScript", "SQL", "AWS" 
    }
    
    all_unique_skills = set(STARTER_KEYWORDS)
    for jd in st.session_state.candidate_jd_list:
        valid_skills = [
            skill.strip() for skill in jd.get('key_skills', []) 
            if isinstance(skill, str) and skill.strip()
        ]
        all_unique_skills.update(valid_skills)
    
    unique_skills_list = sorted(list(all_unique_skills))
    
    if not unique_skills_list:
        unique_skills_list = ["No skills extracted from current JDs"]

    all_jd_data = st.session_state.candidate_jd_list
    # --- End Extraction ---

    # --- Start Filter Form ---
    # This entire block uses st.form, so the filtering only runs on button click
    with st.form(key="jd_filter_form"):
        st.markdown("### Select Filters")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Skills Multiselect
            selected_skills = st.multiselect(
                "Skills Keywords (Select multiple)",
                options=unique_skills_list,
                default=st.session_state.get('last_selected_skills', []),
                key="candidate_filter_skills_multiselect", 
                help="Select one or more skills. JDs containing ANY of the selected skills will be shown."
            )
            
        with col2:
            # Job Type Selectbox
            selected_job_type = st.selectbox(
                "Job Type",
                options=["All Job Types"] + unique_job_types,
                index=0, 
                key="filter_job_type_select"
            )
            
        with col3:
            # Role Title Selectbox
            selected_role = st.selectbox(
                "Role Title",
                options=["All Roles"] + unique_roles,
                index=0, 
                key="filter_role_select"
            )

        # Apply Button (The trigger for the filtering logic)
        apply_filters_button = st.form_submit_button("✅ Apply Filters", type="primary", use_container_width=True)

    # --- Start Filtering Logic (Inside the if apply_filters_button block) ---
    if apply_filters_button:
        
        # Save last selected skills for persistence/re-display if needed
        st.session_state.last_selected_skills = selected_skills

        filtered_jds = []
        
        # Process skills input to lowercase for matching
        selected_skills_lower = [k.strip().lower() for k in selected_skills]
        
        for jd in all_jd_data:
            jd_role = jd.get('role', 'General Analyst')
            jd_job_type = jd.get('job_type', 'Full-time')
            # Ensure JD key skills are also clean and lowercased
            jd_key_skills = [
                s.lower() for s in jd.get('key_skills', []) 
                if isinstance(s, str) and s.strip()
            ]
            
            # 1. Role Filter
            role_match = (selected_role == "All Roles") or (selected_role == jd_role)
            
            # 2. Job Type Filter
            job_type_match = (selected_job_type == "All Job Types") or (selected_job_type == jd_job_type)
            
            # 3. Skills Filter (Match if no skills are selected OR if ANY selected skill is in the JD's key_skills list)
            skill_match = True
            if selected_skills_lower:
                # Check if there is any intersection between selected skills and JD's key skills
                if not any(k in jd_key_skills for k in selected_skills_lower):
                    skill_match = False
            
            # Final Match
            if role_match and job_type_match and skill_match:
                filtered_jds.append(jd)
                
        # Store the filtered list in session state for display persistence
        st.session_state.filtered_jds_display = filtered_jds
        st.success(f"Filter applied! Found {len(filtered_jds)} matching Job Descriptions.")

    # --- Display Results (Always display the last calculated list or an empty list) ---
    st.markdown("---")
    
    # Initialize display list if not set (first run or when JDs are empty)
    if 'filtered_jds_display' not in st.session_state:
        st.session_state.filtered_jds_display = []
        
    filtered_jds = st.session_state.filtered_jds_display
    
    st.subheader(f"Matching Job Descriptions ({len(filtered_jds)} found)")
    
    if filtered_jds:
        display_data = []
        for jd in filtered_jds:
            display_data.append({
                "Job Description Title": jd['name'].replace("--- Simulated JD for: ", ""),
                "Role": jd.get('role', 'N/A'),
                "Job Type": jd.get('job_type', 'N/A'),
                "Key Skills": ", ".join(jd.get('key_skills', ['N/A'])[:5]) + "...",
            })
            
        st.dataframe(display_data, use_container_width=True)

        st.markdown("##### Detailed View")
        for idx, jd in enumerate(filtered_jds, 1):
            with st.expander(f"JD {idx}: {jd['name'].replace('--- Simulated JD for: ', '')} - ({jd.get('role', 'N/A')})"):
                st.markdown(f"**Job Type:** {jd.get('job_type', 'N/A')}")
                st.markdown(f"**Extracted Skills:** {', '.join(jd.get('key_skills', ['N/A']))}")
                st.markdown("---")
                st.text(jd['content'])
    elif st.session_state.candidate_jd_list and apply_filters_button:
        st.info("No Job Descriptions match the selected criteria. Try broadening your filter selections.")
    elif st.session_state.candidate_jd_list and not apply_filters_button:
        st.info("Use the filters above and click **'Apply Filters'** to view matching Job Descriptions.")
        tab_cv_mgmt, tab_parsing, tab_jd_mgmt, tab_batch_match, tab_filter_jd, tab_chatbot, tab_interview_prep = st.tabs([
        "✍️ CV Management", 
        "📄 Resume Parsing", 
        "📚 JD Management", 
        "🎯 Batch JD Match",
        "🔍 Filter JD",
        "💬 Resume/JD Chatbot (Q&A)", # MOVED TO END
        "❓ Interview Prep"            # MOVED TO END
    ])
def candidate_dashboard():
    st.header("👩‍🎓 Candidate Dashboard")
    st.markdown("Welcome! Use the tabs below to manage your CV and access AI preparation tools.")

    # --- MODIFIED NAVIGATION BLOCK ---
    nav_col, _ = st.columns([1, 1]) 

    with nav_col:
        if st.button("🚪 Log Out", key="candidate_logout_btn", use_container_width=True):
            go_to("login") 
    # --- END MODIFIED NAVIGATION BLOCK ---
    
    # Sidebar for Status Only
    with st.sidebar:
        st.header("Resume/CV Status")
        
        # Check if a resume is currently loaded into the main parsing variables
        if st.session_state.parsed.get("name"):
            st.success(f"Currently loaded: **{st.session_state.parsed['name']}**")
        elif st.session_state.full_text:
            st.warning("Resume content is loaded, but parsing may have errors.")
        else:
            st.info("Please upload a file or use the CV builder in 'CV Management' to begin.")

    # Main Content Tabs (REARRANGED TABS HERE)
    tab_cv_mgmt, tab_parsing, tab_jd_mgmt, tab_batch_match, tab_filter_jd, tab_chatbot, tab_interview_prep = st.tabs([
        "✍️ CV Management", 
        "📄 Resume Parsing", 
        "📚 JD Management", 
        "🎯 Batch JD Match",
        "🔍 Filter JD",
        "💬 Resume/JD Chatbot (Q&A)", # MOVED TO END
        "❓ Interview Prep"            # MOVED TO END
    ])
    
    is_resume_parsed = bool(st.session_state.get('parsed', {}).get('name')) or bool(st.session_state.get('full_text'))
    
    # --- TAB 0: CV Management ---
    with tab_cv_mgmt:
        cv_management_tab_content()

    # --- TAB 1 (Now tab_parsing): Resume Parsing (MODIFIED: Added Paste Your CV option) ---
    with tab_parsing:
        st.header("Resume Upload and Parsing")
        
        # 1. Input Method Selection
        input_method = st.radio(
            "Select Input Method",
            ["Upload File", "Paste Text"],
            key="parsing_input_method"
        )
        
        st.markdown("---")

        # --- A. Upload File Method (UPDATED FILE TYPES HERE) ---
        if input_method == "Upload File":
            st.markdown("### 1. Upload Resume File") 
            
            # 🚨 File types expanded here
            uploaded_file = st.file_uploader( 
                "Choose PDF, DOCX, TXT, JSON, MD, CSV, XLSX file", 
                type=["pdf", "docx", "txt", "json", "md", "csv", "xlsx", "markdown", "rtf"], 
                accept_multiple_files=False, 
                key='candidate_file_upload_main'
            )
            
            st.markdown(
                """
                <div style='font-size: 10px; color: grey;'>
                Supported File Types: PDF, DOCX, TXT, JSON, MARKDOWN, CSV, XLSX, RTF
                </div>
                """, 
                unsafe_allow_html=True
            )
            st.markdown("---")

            # --- File Management Logic ---
            if uploaded_file is not None:
                # Only store the single uploaded file if it's new
                if not st.session_state.candidate_uploaded_resumes or st.session_state.candidate_uploaded_resumes[0].name != uploaded_file.name:
                    st.session_state.candidate_uploaded_resumes = [uploaded_file] 
                    st.session_state.pasted_cv_text = "" # Clear pasted text
                    st.toast("Resume file uploaded successfully.")
            elif st.session_state.candidate_uploaded_resumes and uploaded_file is None:
                # Case where the file is removed from the uploader
                st.session_state.candidate_uploaded_resumes = []
                st.session_state.parsed = {}
                st.session_state.full_text = ""
                st.toast("Upload cleared.")
            
            file_to_parse = st.session_state.candidate_uploaded_resumes[0] if st.session_state.candidate_uploaded_resumes else None
            
            st.markdown("### 2. Parse Uploaded File")
            
            if file_to_parse:
                if st.button(f"Parse and Load: **{file_to_parse.name}**", use_container_width=True):
                    with st.spinner(f"Parsing {file_to_parse.name}..."):
                        result = parse_and_store_resume(file_to_parse, file_name_key='single_resume_candidate', source_type='file')
                        
                        if "error" not in result:
                            st.session_state.parsed = result['parsed']
                            st.session_state.full_text = result['full_text']
                            st.session_state.excel_data = result['excel_data'] 
                            st.session_state.parsed['name'] = result['name'] 
                            clear_interview_state()
                            st.success(f"✅ Successfully loaded and parsed **{result['name']}**.")
                            st.info("View, edit, and download the parsed data in the **CV Management** tab.") 
                        else:
                            st.error(f"Parsing failed for {file_to_parse.name}: {result['error']}")
                            st.session_state.parsed = {"error": result['error'], "name": result['name']}
                            st.session_state.full_text = result['full_text'] or ""
            else:
                st.info("No resume file is currently uploaded. Please upload a file above.")

        # --- B. Paste Text Method (NEW) ---
        else: # input_method == "Paste Text"
            st.markdown("### 1. Paste Your CV Text")
            
            pasted_text = st.text_area(
                "Copy and paste your entire CV or resume text here.",
                value=st.session_state.get('pasted_cv_text', ''),
                height=300,
                key='pasted_cv_text_input'
            )
            st.session_state.pasted_cv_text = pasted_text # Update session state immediately
            
            st.markdown("---")
            st.markdown("### 2. Parse Pasted Text")
            
            if pasted_text.strip():
                if st.button("Parse and Load Pasted Text", use_container_width=True):
                    with st.spinner("Parsing pasted text..."):
                        # Clear file upload state
                        st.session_state.candidate_uploaded_resumes = []
                        
                        result = parse_and_store_resume(pasted_text, file_name_key='single_resume_candidate', source_type='text')
                        
                        if "error" not in result:
                            st.session_state.parsed = result['parsed']
                            st.session_state.full_text = result['full_text']
                            st.session_state.excel_data = result['excel_data'] 
                            st.session_state.parsed['name'] = result['name'] 
                            clear_interview_state()
                            st.success(f"✅ Successfully loaded and parsed **{result['name']}**.")
                            st.info("View, edit, and download the parsed data in the **CV Management** tab.") 
                        else:
                            st.error(f"Parsing failed: {result['error']}")
                            st.session_state.parsed = {"error": result['error'], "name": result['name']}
                            st.session_state.full_text = result['full_text'] or ""
            else:
                st.info("Please paste your CV text into the box above.")

    # --- TAB 2 (Now tab_jd_mgmt): JD Management (Candidate) ---
    with tab_jd_mgmt:
        st.header("📚 Manage Job Descriptions for Matching")
        st.markdown("Add multiple JDs here to compare your resume against them in the next tabs.")
        
        if "candidate_jd_list" not in st.session_state:
             st.session_state.candidate_jd_list = []
        
        jd_type = st.radio("Select JD Type", ["Single JD", "Multiple JD"], key="jd_type_candidate")
        st.markdown("### Add JD by:")
        
        method = st.radio("Choose Method", ["Upload File", "Paste Text", "LinkedIn URL"], key="jd_add_method_candidate") 

        # URL
        if method == "LinkedIn URL":
            url_list = st.text_area(
                "Enter one or more URLs (comma separated)" if jd_type == "Multiple JD" else "Enter URL", key="url_list_candidate"
            )
            if st.button("Add JD(s) from URL", key="add_jd_url_btn_candidate"):
                if url_list:
                    urls = [u.strip() for u in url_list.split(",")] if jd_type == "Multiple JD" else [url_list.strip()]
                    
                    count = 0
                    for url in urls:
                        if not url: continue
                        
                        with st.spinner(f"Attempting JD extraction and metadata analysis for: {url}"):
                            jd_text = extract_jd_from_linkedin_url(url)
                            metadata = extract_jd_metadata(jd_text) # NEW METADATA EXTRACTION
                        
                        name_base = url.split('/jobs/view/')[-1].split('/')[0] if '/jobs/view/' in url else f"URL {count+1}"
                        # CRITICAL: Added explicit JD naming convention for LinkedIn URLs in Candidate JD list
                        name = f"JD from URL: {name_base}" 
                        if name in [item['name'] for item in st.session_state.candidate_jd_list]:
                            name = f"JD from URL: {name_base} ({len(st.session_state.candidate_jd_list) + 1})" 

                        st.session_state.candidate_jd_list.append({"name": name, "content": jd_text, **metadata}) # ADD METADATA
                        
                        if not jd_text.startswith("[Error"):
                            count += 1
                                
                    if count > 0:
                        st.success(f"✅ {count} JD(s) added successfully! Check the display below for the extracted content.")
                    else:
                        st.error("No JDs were added successfully.")


        # Paste Text
        elif method == "Paste Text":
            text_list = st.text_area(
                "Paste one or more JD texts (separate by '---')" if jd_type == "Multiple JD" else "Paste JD text here", key="text_list_candidate"
            )
            if st.button("Add JD(s) from Text", key="add_jd_text_btn_candidate"):
                if text_list:
                    texts = [t.strip() for t in text_list.split("---")] if jd_type == "Multiple JD" else [text_list.strip()]
                    for i, text in enumerate(texts):
                         if text:
                            name_base = text.splitlines()[0].strip()
                            if len(name_base) > 30: name_base = f"{name_base[:27]}..."
                            if not name_base: name_base = f"Pasted JD {len(st.session_state.candidate_jd_list) + i + 1}"
                            
                            metadata = extract_jd_metadata(text) # NEW METADATA EXTRACTION
                            st.session_state.candidate_jd_list.append({"name": name_base, "content": text, **metadata}) # ADD METADATA
                    st.success(f"✅ {len(texts)} JD(s) added successfully!")

        # Upload File
        elif method == "Upload File":
            uploaded_files = st.file_uploader(
                "Upload JD file(s)",
                type=["pdf", "txt", "docx"],
                accept_multiple_files=(jd_type == "Multiple JD"), # Dynamically set
                key="jd_file_uploader_candidate"
            )
            if st.button("Add JD(s) from File", key="add_jd_file_btn_candidate"):
                # CRITICAL FIX: Ensure 'files_to_process' is always a list of single UploadedFile objects
                if uploaded_files is None:
                    st.warning("Please upload file(s).")
                    
                files_to_process = uploaded_files if isinstance(uploaded_files, list) else ([uploaded_files] if uploaded_files else [])
                
                count = 0
                for file in files_to_process:
                    if file:
                        temp_dir = tempfile.mkdtemp()
                        temp_path = os.path.join(temp_dir, file.name)
                        with open(temp_path, "wb") as f:
                            f.write(file.getbuffer())
                            
                        file_type = get_file_type(temp_path)
                        jd_text = extract_content(file_type, temp_path)
                        
                        if not jd_text.startswith("Error"):
                            metadata = extract_jd_metadata(jd_text) # NEW METADATA EXTRACTION
                            st.session_state.candidate_jd_list.append({"name": file.name, "content": jd_text, **metadata}) # ADD METADATA
                            count += 1
                        else:
                            st.error(f"Error extracting content from {file.name}: {jd_text}")
                            
                if count > 0:
                    st.success(f"✅ {count} JD(s) added successfully!")
                elif uploaded_files:
                    st.error("No valid JD files were uploaded or content extraction failed.")


        # Display Added JDs
        if st.session_state.candidate_jd_list:
            
            col_display_header, col_clear_button = st.columns([3, 1])
            
            with col_display_header:
                st.markdown("### ✅ Current JDs Added:")
                
            with col_clear_button:
                if st.button("🗑️ Clear All JDs", key="clear_jds_candidate", use_container_width=True, help="Removes all currently loaded JDs."):
                    st.session_state.candidate_jd_list = []
                    st.session_state.candidate_match_results = [] 
                    # Also clear filter display
                    st.session_state.filtered_jds_display = [] 
                    st.success("All JDs and associated match results have been cleared.")
                    st.rerun() 

            for idx, jd_item in enumerate(st.session_state.candidate_jd_list, 1):
                title = jd_item['name']
                display_title = title.replace("--- Simulated JD for: ", "")
                with st.expander(f"JD {idx}: {display_title} | Role: {jd_item.get('role', 'N/A')}"):
                    st.markdown(f"**Job Type:** {jd_item.get('job_type', 'N/A')} | **Key Skills:** {', '.join(jd_item.get('key_skills', ['N/A']))}") # ADDED METADATA DISPLAY
                    st.markdown("---")
                    st.text(jd_item['content'])
        else:
            st.info("No Job Descriptions added yet.")

    # --- TAB 3 (Now tab_batch_match): Batch JD Match (Candidate) ---
    with tab_batch_match:
        st.header("🎯 Batch JD Match: Best Matches")
        st.markdown("Compare your current resume against all saved job descriptions.")

        if not is_resume_parsed:
            st.warning("Please **upload and parse your resume** in the 'Resume Parsing' tab or **build your CV** in the 'CV Management' tab first.")
        
        elif not st.session_state.candidate_jd_list:
            st.error("Please **add Job Descriptions** in the 'JD Management' tab (Tab 4) before running batch analysis.")
            
        elif not GROQ_API_KEY:
             st.error("Cannot use JD Match: GROQ_API_KEY is not configured.")
             
        else:
            if "candidate_match_results" not in st.session_state:
                st.session_state.candidate_match_results = []

            # 1. Get all available JD names
            all_jd_names = [item['name'] for item in st.session_state.candidate_jd_list]
            
            # 2. Add multiselect widget
            selected_jd_names = st.multiselect(
                "Select Job Descriptions to Match Against",
                options=all_jd_names,
                default=all_jd_names, # Default to selecting all JDs
                key='candidate_batch_jd_select'
            )
            
            # 3. Filter the list of JDs based on selection
            jds_to_match = [
                jd_item for jd_item in st.session_state.candidate_jd_list 
                if jd_item['name'] in selected_jd_names
            ]
            
            if st.button(f"Run Match Analysis on {len(jds_to_match)} Selected JD(s)"):
                st.session_state.candidate_match_results = []
                
                if not jds_to_match:
                    st.warning("Please select at least one Job Description to run the analysis.")
                    
                else:
                    resume_name = st.session_state.parsed.get('name', 'Uploaded Resume')
                    parsed_json = st.session_state.parsed
                    results_with_score = []

                    with st.spinner(f"Matching {resume_name}'s resume against {len(jds_to_match)} selected JD(s)..."):
                        
                        # Loop over jds_to_match
                        for jd_item in jds_to_match:
                            
                            jd_name = jd_item['name']
                            jd_content = jd_item['content']

                            try:
                                fit_output = evaluate_jd_fit(jd_content, parsed_json)
                                
                                overall_score_match = re.search(r'Overall Fit Score:\s*[^\d]*(\d+)\s*/10', fit_output, re.IGNORECASE)
                                section_analysis_match = re.search(
                                    r'--- Section Match Analysis ---\s*(.*?)\s*Strengths/Matches:', 
                                    fit_output, re.DOTALL
                                )

                                skills_percent, experience_percent, education_percent = 'N/A', 'N/A', 'N/A'
                                
                                if section_analysis_match:
                                    section_text = section_analysis_match.group(1)
                                    skills_match = re.search(r'Skills Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                                    experience_match = re.search(r'Experience Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                                    education_match = re.search(r'Education Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                                    
                                    if skills_match: skills_percent = skills_match.group(1)
                                    if experience_match: experience_percent = experience_match.group(1)
                                    if education_match: education_percent = education_match.group(1)
                                
                                overall_score = overall_score_match.group(1) if overall_score_match else 'N/A'

                                results_with_score.append({
                                    "jd_name": jd_name,
                                    "overall_score": overall_score,
                                    "numeric_score": int(overall_score) if overall_score.isdigit() else -1, # Added for sorting/ranking
                                    "skills_percent": skills_percent,
                                    "experience_percent": experience_percent, 
                                    "education_percent": education_percent,   
                                    "full_analysis": fit_output
                                })
                            except Exception as e:
                                results_with_score.append({
                                    "jd_name": jd_name,
                                    "overall_score": "Error",
                                    "numeric_score": -1, # Set a low score for errors
                                    "skills_percent": "Error",
                                    "experience_percent": "Error", 
                                    "education_percent": "Error",   
                                    "full_analysis": f"Error running analysis for {jd_name}: {e}\n{traceback.format_exc()}"
                                })
                                
                        # --- NEW RANKING LOGIC ---
                        # 1. Sort by numeric_score (highest first)
                        results_with_score.sort(key=lambda x: x['numeric_score'], reverse=True)
                        
                        # 2. Assign Rank (handle ties)
                        current_rank = 1
                        current_score = -1 
                        
                        for i, item in enumerate(results_with_score):
                            if item['numeric_score'] > current_score:
                                current_rank = i + 1
                                current_score = item['numeric_score']
                            
                            item['rank'] = current_rank
                            # Remove the temporary numeric_score field
                            del item['numeric_score'] 
                            
                        st.session_state.candidate_match_results = results_with_score
                        # --- END NEW RANKING LOGIC ---
                        
                        st.success("Batch analysis complete!")


            # 3. Display Results (UPDATED TO INCLUDE RANK)
            if st.session_state.get('candidate_match_results'):
                st.markdown("#### Match Results for Your Resume")
                results_df = st.session_state.candidate_match_results
                
                display_data = []
                for item in results_df:
                    # Also include extracted JD metadata for a richer view
                    
                    # Find the full JD item to get the metadata
                    full_jd_item = next((jd for jd in st.session_state.candidate_jd_list if jd['name'] == item['jd_name']), {})
                    
                    display_data.append({
                        # 🚨 ADDED RANK COLUMN
                        "Rank": item.get("rank", "N/A"),
                        "Job Description (Ranked)": item["jd_name"].replace("--- Simulated JD for: ", ""),
                        "Role": full_jd_item.get('role', 'N/A'), # Added Role
                        "Job Type": full_jd_item.get('job_type', 'N/A'), # Added Job Type
                        "Fit Score (out of 10)": item["overall_score"],
                        "Skills (%)": item.get("skills_percent", "N/A"),
                        "Experience (%)": item.get("experience_percent", "N/A"), 
                        "Education (%)": item.get("education_percent", "N/A"),   
                    })

                st.dataframe(display_data, use_container_width=True)

                st.markdown("##### Detailed Reports")
                for item in results_df:
                    # UPDATED HEADER TO INCLUDE RANK
                    rank_display = f"Rank {item.get('rank', 'N/A')} | "
                    header_text = f"{rank_display}Report for **{item['jd_name'].replace('--- Simulated JD for: ', '')}** (Score: **{item['overall_score']}/10** | S: **{item.get('skills_percent', 'N/A')}%** | E: **{item.get('experience_percent', 'N/A')}%** | Edu: **{item.get('education_percent', 'N/A')}%**)"
                    with st.expander(header_text):
                        st.markdown(item['full_analysis'])

    # --- TAB 4 (Now tab_filter_jd): Filter JD (NEW) ---
    with tab_filter_jd:
        filter_jd_tab_content()

    # --- TAB 5 (Now tab_chatbot): Resume Chatbot (Q&A) (MOVED) ---
    with tab_chatbot:
        st.header("Resume/JD Chatbot (Q&A) 💬")
        
        # --- NESTED TABS ---
        sub_tab_resume, sub_tab_jd = st.tabs([
            "👤 Chat about Your Resume",
            "📄 Chat about Saved JDs"
        ])
        
        # --- 5A. RESUME CHATBOT CONTENT ---
        with sub_tab_resume:
            st.markdown("### Ask any question about the currently loaded resume.")
            if not is_resume_parsed:
                st.warning("Please upload and parse a resume in the 'Resume Parsing' tab or use the 'CV Management' tab first.")
            elif "error" in st.session_state.parsed:
                 st.error("Cannot use Resume Chatbot: Resume data has parsing errors.")
            elif not GROQ_API_KEY:
                 st.error("Cannot use Chatbot: GROQ_API_KEY is not configured.")
            else:
                if 'qa_answer_resume' not in st.session_state: st.session_state.qa_answer_resume = ""
                
                question = st.text_input(
                    "Your Question (about Resume)", 
                    placeholder="e.g., What are the candidate's key skills?",
                    key="resume_qa_question"
                )
                
                if st.button("Get Answer (Resume)", key="qa_btn_resume"):
                    with st.spinner("Generating answer..."):
                        try:
                            answer = qa_on_resume(question)
                            st.session_state.qa_answer_resume = answer
                        except Exception as e:
                            st.error(f"Error during Resume Q&A: {e}")
                            st.session_state.qa_answer_resume = "Could not generate an answer."

                if st.session_state.get('qa_answer_resume'):
                    st.text_area("Answer (Resume)", st.session_state.qa_answer_resume, height=150)
        
        # --- 5B. JD CHATBOT CONTENT (NEW SUBTAB) ---
        with sub_tab_jd:
            st.markdown("### Ask any question about a saved Job Description.")
            
            if not st.session_state.candidate_jd_list:
                st.warning("Please add Job Descriptions in the 'JD Management' tab (Tab 4) first.")
            elif not GROQ_API_KEY:
                 st.error("Cannot use JD Chatbot: GROQ_API_KEY is not configured.")
            else:
                if 'qa_answer_jd' not in st.session_state: st.session_state.qa_answer_jd = ""

                # 1. JD Selection
                jd_names = [jd['name'] for jd in st.session_state.candidate_jd_list]
                selected_jd_name = st.selectbox(
                    "Select Job Description to Query",
                    options=jd_names,
                    key="jd_qa_select"
                )
                
                # 2. Question Input
                question = st.text_input(
                    "Your Question (about JD)", 
                    placeholder="e.g., What is the minimum experience required for this role?",
                    key="jd_qa_question"
                )
                
                # 3. Get Answer Button
                if st.button("Get Answer (JD)", key="qa_btn_jd"):
                    if selected_jd_name and question.strip():
                        with st.spinner(f"Generating answer for {selected_jd_name}..."):
                            try:
                                answer = qa_on_jd(question, selected_jd_name)
                                st.session_state.qa_answer_jd = answer
                            except Exception as e:
                                st.error(f"Error during JD Q&A: {e}")
                                st.session_state.qa_answer_jd = "Could not generate an answer."
                    else:
                        st.error("Please select a JD and enter a question.")

                # 4. Answer Output
                if st.session_state.get('qa_answer_jd'):
                    st.text_area("Answer (JD)", st.session_state.qa_answer_jd, height=150)


    # --- TAB 6 (Now tab_interview_prep): Interview Prep (MOVED) ---
    with tab_interview_prep:
        st.header("Interview Preparation Tools")
        if not is_resume_parsed or "error" in st.session_state.parsed:
            st.warning("Please upload and successfully parse a resume first.")
        elif not GROQ_API_KEY:
             st.error("Cannot use Interview Prep: GROQ_API_KEY is not configured.")
        else:
            if 'iq_output' not in st.session_state: st.session_state.iq_output = ""
            if 'interview_qa' not in st.session_state: st.session_state.interview_qa = [] 
            if 'evaluation_report' not in st.session_state: st.session_state.evaluation_report = "" 
            
            st.subheader("1. Generate Interview Questions")
            
            section_choice = st.selectbox(
                "Select Section", 
                question_section_options, 
                key='iq_section_c',
                on_change=clear_interview_state 
            )
            
            if st.button("Generate Interview Questions", key='iq_btn_c'):
                with st.spinner("Generating questions..."):
                    try:
                        raw_questions_response = generate_interview_questions(st.session_state.parsed, section_choice)
                        st.session_state.iq_output = raw_questions_response
                        
                        st.session_state.interview_qa = [] 
                        st.session_state.evaluation_report = "" 
                        
                        q_list = []
                        current_level = ""
                        for line in raw_questions_response.splitlines():
                            line = line.strip()
                            if line.startswith('[') and line.endswith(']'):
                                current_level = line.strip('[]')
                            elif line.lower().startswith('q') and ':' in line:
                                question_text = line[line.find(':') + 1:].strip()
                                q_list.append({
                                    "question": f"({current_level}) {question_text}",
                                    "answer": "", 
                                    "level": current_level
                                })
                                
                        st.session_state.interview_qa = q_list
                        
                        st.success(f"Generated {len(q_list)} questions based on your **{section_choice}** section.")
                        
                    except Exception as e:
                        st.error(f"Error generating questions: {e}")
                        st.session_state.iq_output = "Error generating questions."
                        st.session_state.interview_qa = []

            if st.session_state.get('interview_qa'):
                st.markdown("---")
                st.subheader("2. Practice and Record Answers")
                
                with st.form("interview_practice_form"):
                    
                    for i, qa_item in enumerate(st.session_state.interview_qa):
                        st.markdown(f"**Question {i+1}:** {qa_item['question']}")
                        
                        answer = st.text_area(
                            f"Your Answer for Q{i+1}", 
                            value=st.session_state.interview_qa[i]['answer'], 
                            height=100,
                            key=f'answer_q_{i}',
                            label_visibility='collapsed'
                        )
                        st.session_state.interview_qa[i]['answer'] = answer 
                        st.markdown("---") 
                        
                    submit_button = st.form_submit_button("Submit & Evaluate Answers", use_container_width=True)

                    if submit_button:
                        
                        if all(item['answer'].strip() for item in st.session_state.interview_qa):
                            with st.spinner("Sending answers to AI Evaluator..."):
                                try:
                                    report = evaluate_interview_answers(
                                        st.session_state.interview_qa,
                                        st.session_state.parsed
                                    )
                                    st.session_state.evaluation_report = report
                                    st.success("Evaluation complete! See the report below.")
                                except Exception as e:
                                    st.error(f"Evaluation failed: {e}")
                                    st.session_state.evaluation_report = f"Evaluation failed: {e}\n{traceback.format_exc()}"
                        else:
                            st.error("Please answer all generated questions before submitting.")
                
                if st.session_state.get('evaluation_report'):
                    st.markdown("---")
                    st.subheader("3. AI Evaluation Report")
                    st.markdown(st.session_state.evaluation_report)
                    def hiring_dashboard():
    st.header("🏢 Hiring Company Dashboard")
    st.write("Manage job postings and view candidate applications. (Placeholder for future features)")
    
    # --- MODIFIED NAVIGATION BLOCK (MODIFIED) ---
    nav_col, _ = st.columns([1, 1]) 

    with nav_col:
        if st.button("🚪 Log Out", key="hiring_logout_btn", use_container_width=True):
            go_to("login") 
    # --- END MODIFIED NAVIGATION BLOCK ---

# -------------------------
# Main App Initialization
# -------------------------
def main():
    st.set_page_config(layout="wide", page_title="PragyanAI Job Portal")

    # --- Session State Initialization ---
    if 'page' not in st.session_state: st.session_state.page = "login"
    
    # Initialize session state for AI features (Defensive Initialization)
    if 'parsed' not in st.session_state: st.session_state.parsed = {}
    if 'full_text' not in st.session_state: st.session_state.full_text = ""
    if 'excel_data' not in st.session_state: st.session_state.excel_data = None
    
    # Chatbot/Q&A answers (Modified to distinguish Resume and JD)
    if 'qa_answer_resume' not in st.session_state: st.session_state.qa_answer_resume = ""
    if 'qa_answer_jd' not in st.session_state: st.session_state.qa_answer_jd = ""
    
    if 'iq_output' not in st.session_state: st.session_state.iq_output = ""
    if 'jd_fit_output' not in st.session_state: st.session_state.jd_fit_output = ""
        
        # Admin Dashboard specific lists
    if 'admin_jd_list' not in st.session_state: st.session_state.admin_jd_list = [] 
    if 'resumes_to_analyze' not in st.session_state: st.session_state.resumes_to_analyze = [] 
    if 'admin_match_results' not in st.session_state: st.session_state.admin_match_results = []
    if 'resume_statuses' not in st.session_state: st.session_state.resume_statuses = {} 
        
        # Vendor State Init
    if 'vendors' not in st.session_state: st.session_state.vendors = []
    if 'vendor_statuses' not in st.session_state: st.session_state.vendor_statuses = {}
        
        # Candidate Dashboard specific lists
    # NOTE: These JD items now store content, name, role, job_type, and key_skills
    if 'candidate_jd_list' not in st.session_state: st.session_state.candidate_jd_list = []
    if 'candidate_match_results' not in st.session_state: st.session_state.candidate_match_results = []
    
    # Resume Parsing Upload State
    if 'candidate_uploaded_resumes' not in st.session_state: st.session_state.candidate_uploaded_resumes = []
    
    # NEW: Pasted Text State
    if 'pasted_cv_text' not in st.session_state: st.session_state.pasted_cv_text = "" 
    
    # Interview Prep Q&A State (NEW)
    if 'interview_qa' not in st.session_state: st.session_state.interview_qa = [] 
    if 'evaluation_report' not in st.session_state: st.session_state.evaluation_report = ""
        
    # CV Builder Form State (NEW)
    if "cv_form_data" not in st.session_state: 
        st.session_state.cv_form_data = {
            "name": "", "email": "", "phone": "", "linkedin": "", "github": "",
            "skills": [], "experience": [], "education": [], "certifications": [], 
            "projects": [], "strength": [], "personal_details": ""
        }
    
    # Filter State (NEW)
    if "candidate_filter_skills_multiselect" not in st.session_state:
        st.session_state.candidate_filter_skills_multiselect = []
    if "filtered_jds_display" not in st.session_state:
        st.session_state.filtered_jds_display = []
    if "last_selected_skills" not in st.session_state:
        st.session_state.last_selected_skills = []


    # --- Page Routing ---
    if st.session_state.page == "login":
        login_page()
    elif st.session_state.page == "signup":
        signup_page()
    elif st.session_state.page == "admin_dashboard":
        admin_dashboard()
    elif st.session_state.page == "candidate_dashboard":
        candidate_dashboard()
    elif st.session_state.page == "hiring_dashboard":
        hiring_dashboard()

if __name__ == '__main__':
    main()
