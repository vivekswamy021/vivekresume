import streamlit as st
import streamlit as st
import os
import pdfplumber
import docx
import openpyxl
import json
import tempfile
from groq import Groq
import traceback
import re 
from dotenv import load_dotenv 
from datetime import date 
from streamlit.runtime.uploaded_file_manager import UploadedFile

# -------------------------
# CONFIGURATION & API SETUP (Necessary for standalone functions)
# -------------------------

GROQ_MODEL = "llama-3.1-8b-instant"
# Load environment variables (mocked if running standalone)
load_dotenv()
GROQ_API_KEY = os.getenv('GROQ_API_KEY')

# Initialize Groq Client or Mock Client (Must be present for function definitions)
if not GROQ_API_KEY:
    class MockGroqClient:
        def chat(self):
            class Completions:
                def create(self, **kwargs):
                    raise ValueError("GROQ_API_KEY not set. AI functions disabled.")
            return Completions()
    client = MockGroqClient()
else:
    client = Groq(api_key=GROQ_API_KEY)

# --- Utility Functions (Only necessary ones for Admin) ---

def go_to(page_name):
    """Changes the current page in Streamlit's session state."""
    st.session_state.page = page_name

def get_file_type(file_path):
    """Identifies the file type based on its extension."""
    ext = os.path.splitext(file_path)[1].lower().strip('.')
    if ext == 'pdf': return 'pdf'
    elif ext == 'docx': return 'docx'
    elif ext == 'xlsx': return 'xlsx'
    else: return 'txt' 

def extract_content(file_type, file_path):
    """Extracts text content from various file types (Simplified for admin context)."""
    text = ''
    try:
        if file_type == 'pdf':
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + '\n'
        elif file_type == 'docx':
            doc = docx.Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
        elif file_type == 'xlsx':
            workbook = openpyxl.load_workbook(file_path)
            # Simplified XLSX reading for content
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join([str(c) for c in row if c is not None])
                    if row_text.strip():
                        text += row_text + '\n'
        else:
             with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()

        if not text.strip():
            return f"Error: {file_type.upper()} content extraction failed."
        
        return text
    
    except Exception as e:
        return f"Fatal Extraction Error: Failed to read file content ({file_type}). Error: {e}"


@st.cache_data(show_spinner="Extracting JD metadata...")
def extract_jd_metadata(jd_text):
    """Extracts structured metadata (Role, Job Type, Key Skills) from raw JD text."""
    if not GROQ_API_KEY:
        return {"role": "N/A", "job_type": "N/A", "key_skills": []}

    prompt = f"""Analyze the following Job Description and extract the key metadata.
    
    Job Description:
    {jd_text}
    
    Provide the output strictly as a JSON object with the following three keys:
    1.  **role**: The main job title (e.g., 'Data Scientist', 'Senior Software Engineer'). If not clear, default to 'General Analyst'.
    2.  **job_type**: The employment type (e.g., 'Full-time', 'Contract', 'Internship', 'Remote'). If not clear, default to 'Full-time'.
    3.  **key_skills**: A list of 5 to 10 most critical hard and soft skills required (e.g., ['Python', 'AWS', 'Teamwork', 'SQL']).
    
    Example Output: {{"role": "Software Engineer", "job_type": "Full-time", "key_skills": ["Python", "JavaScript", "React", "AWS", "Agile"]}}
    """
    content = ""
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        content = response.choices[0].message.content.strip()

        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0).strip()
            if json_str.startswith('```json'):
                json_str = json_str[len('```json'):].strip()
            if json_str.endswith('```'):
                json_str = json_str[:-len('```')].strip()
            
            parsed = json.loads(json_str)
        else:
            raise json.JSONDecodeError("Could not isolate a valid JSON structure from LLM response.", content, 0)
        
        return {
            "role": parsed.get("role", "General Analyst"),
            "job_type": parsed.get("job_type", "Full-time"),
            "key_skills": [s.strip() for s in parsed.get("key_skills", []) if isinstance(s, str)]
        }

    except Exception:
        return {"role": "General Analyst (LLM Error)", "job_type": "Full-time (LLM Error)", "key_skills": ["LLM Error", "Fallback"]}


@st.cache_data(show_spinner="Analyzing content with Groq LLM...")
def parse_with_llm(text, return_type='json'):
    """Sends resume text to the LLM for structured information extraction (Simplified for admin context)."""
    if text.startswith("Error") or not GROQ_API_KEY:
        return {"error": "Parsing error or API key missing.", "raw_output": ""}

    prompt = f"""Extract the following information from the resume in structured JSON.
    - Name, - Email, - Phone, - Skills, - Education, 
    - Experience, - Certifications, 
    - Projects, - Strength, 
    - Personal Details, - Github, - LinkedIn
    
    Also, provide a key called **'summary'** which is a single, brief paragraph (3-4 sentences max) summarizing the candidate's career highlights and most relevant skills.
    
    Resume Text: {text}
    
    Provide the output strictly as a JSON object.
    """
    content = ""
    parsed = {}
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2
        )
        content = response.choices[0].message.content.strip()
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0).strip()
            json_str = json_str.replace('```json', '').replace('```', '').strip()
            parsed = json.loads(json_str)
        else:
            raise json.JSONDecodeError("Could not isolate a valid JSON structure.", content, 0)
    except Exception as e:
        parsed = {"error": f"LLM error: {e}", "raw_output": content}

    return parsed


def extract_jd_from_linkedin_url(url: str) -> str:
    """Simulates JD content extraction from a LinkedIn URL."""
    if "linkedin.com/jobs/" not in url:
         return f"[Error: Not a valid LinkedIn Job URL format: {url}]"

    job_title = "Data Scientist"
    try:
        match = re.search(r'/jobs/view/([^/]+)', url) or re.search(r'/jobs/(\w+)', url)
        if match:
            job_title = match.group(1).split('?')[0].replace('-', ' ').title()
    except:
        pass
    
    return f"""
    --- Simulated JD for: {job_title} ---
    **Company:** Quantum Analytics Inc.
    **Role:** {job_title}
    **Responsibilities:** Develop and implement machine learning models. Clean and analyze large datasets using Python/R and SQL. Deploy models into production environments.
    **Requirements:** MS/PhD in Computer Science, Statistics, or a quantitative field. 3+ years experience. Expertise in Python (Pandas, Scikit-learn). Experience with cloud platforms (AWS, Azure, or GCP).
    --- End Simulated JD ---
    """.strip()


def evaluate_jd_fit(job_description, parsed_json):
    """Evaluates how well a resume fits a given job description, including section-wise scores."""
    if not GROQ_API_KEY or "error" in parsed_json: return "AI Evaluation Disabled or resume parsing failed."
    
    relevant_resume_data = {
        'Skills': parsed_json.get('skills', 'Not found or empty'),
        'Experience': parsed_json.get('experience', 'Not found or empty'),
        'Education': parsed_json.get('education', 'Not found or empty'),
    }
    resume_summary = json.dumps(relevant_resume_data, indent=2)

    prompt = f"""Evaluate how well the following resume content matches the provided job description.
    Job Description: {job_description}
    Resume Sections for Analysis: {resume_summary}
    Provide a detailed evaluation structured as follows:
    1.  **Overall Fit Score:** A score out of 10.
    2.  **Section Match Percentages:** A percentage score for the match in the key sections (Skills, Experience, Education).
    ... [rest of prompt truncated for brevity]
    Format the output strictly as follows:
    Overall Fit Score: [Score]/10
    
    --- Section Match Analysis ---
    Skills Match: [XX]%
    Experience Match: [YY]%
    Education Match: [ZZ]%
    
    Strengths/Matches:
    - Point 1
    ...
    """

    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.3
    )
    return response.choices[0].message.content.strip()


def parse_and_store_resume(file_input, file_name_key='default', source_type='file'):
    """Handles file/text input, parsing, and stores results (Simplified for admin context)."""
    text = None
    file_name = f"Pasted Text ({date.today().strftime('%Y-%m-%d')})"

    if source_type == 'file':
        # Admin parsing always runs from a temp file for uploaded files
        if not isinstance(file_input, UploadedFile):
            return {"error": "Invalid file input type passed to parser.", "full_text": ""}

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, file_input.name) 
        with open(temp_path, "wb") as f:
            f.write(file_input.getbuffer()) 

        file_type = get_file_type(temp_path)
        text = extract_content(file_type, temp_path)
        file_name = file_input.name.split('.')[0]
    
    elif source_type == 'text':
        text = file_input
        file_name = f"Pasted Text ({date.today().strftime('%Y-%m-%d')})"
        
    if text.startswith("Error"):
        return {"error": text, "full_text": text, "name": file_name}

    parsed = parse_with_llm(text, return_type='json')
    
    if "error" in parsed:
        return {"error": parsed.get('error', 'Unknown parsing error'), "full_text": text, "name": file_name}
    
    final_name = parsed.get('name', file_name)

    return {
        "parsed": parsed,
        "full_text": text,
        "excel_data": None, # Removed Excel logic for brevity in this isolated block
        "name": final_name
    }


def update_resume_metadata(resume_name, new_status, applied_jd, submitted_date, resume_list_index):
    """Callback function to update the status and metadata of a specific resume."""
    # Update Status
    st.session_state.resume_statuses[resume_name] = new_status
    
    # Update Metadata (Applied JD and Date)
    if 0 <= resume_list_index < len(st.session_state.resumes_to_analyze):
        st.session_state.resumes_to_analyze[resume_list_index]['applied_jd'] = applied_jd
        st.session_state.resumes_to_analyze[resume_list_index]['submitted_date'] = submitted_date
        st.toast(f"Status for **{resume_name}** updated to **{new_status}**.")
    else:
        st.error(f"Error: Could not find resume index {resume_list_index} for metadata update.")
        
# --- Approval Tab Content Functions (Used within admin_dashboard) ---

def candidate_approval_tab_content():
    st.header("👤 Candidate Approval")
    st.markdown("### Review and Set Status for Submitted Resumes")
    
    if "resumes_to_analyze" not in st.session_state or not st.session_state.resumes_to_analyze:
        st.info("No resumes have been uploaded and parsed in the 'Resume Analysis' tab yet.")
        return
        
    jd_options = [item['name'].replace("--- Simulated JD for: ", "") for item in st.session_state.admin_jd_list]
    jd_options.insert(0, "Select JD") 

    for idx, resume_data in enumerate(st.session_state.resumes_to_analyze):
        resume_name = resume_data['name']
        current_status = st.session_state.resume_statuses.get(resume_name, "Pending")
        
        # --- Extract details from parsed JSON ---
        parsed_data = resume_data.get('parsed', {})
        candidate_email = parsed_data.get('email', 'N/A')
        candidate_phone = parsed_data.get('phone', 'N/A')
        
        # Find highest education/university
        education_list = parsed_data.get('education', [])
        university_info = "N/A"
        if education_list:
            # Simple heuristic: take the first item, often the highest degree or most recent
            university_info = education_list[0] 
            # Trim if too long
            if len(university_info) > 60:
                 university_info = university_info[:57] + "..."

        brief_summary = parsed_data.get('summary', 'AI summary pending or failed during parsing.')
        
        # --- Current Metadata (Used for display and form defaults) ---
        current_applied_jd = resume_data.get('applied_jd', 'N/A (Pending Assignment)')
        current_submitted_date = resume_data.get('submitted_date', date.today().strftime("%Y-%m-%d"))

        # --- Display and Action Block for Individual Candidate ---
        with st.container(border=True):
            st.markdown(f"### **Candidate:** {resume_name} (Status: **{current_status}**)")
            
            # Contact Info & Education
            col_contact, col_education = st.columns(2)
            with col_contact:
                st.markdown(f"**📧 Email:** `{candidate_email}`")
                st.markdown(f"**📱 Phone:** `{candidate_phone}`")
            with col_education:
                st.markdown(f"**🎓 Education:** `{university_info}`")
                st.markdown(f"**Applied JD:** `{current_applied_jd}`")
                
            st.markdown("---")
            st.markdown(f"**Brief Resume Info:** *{brief_summary}*")
            st.markdown("---")
            
            # NEW: JD Selection and Date Input Block (No generic status selector/updater)
            col_jd_select, col_date_input = st.columns([1, 1])
            
            with col_jd_select:
                try:
                    default_value = current_applied_jd if current_applied_jd != "N/A (Pending Assignment)" else "Select JD"
                    jd_default_index = jd_options.index(default_value)
                except ValueError:
                    jd_default_index = 0
                    
                new_applied_jd = st.selectbox(
                    "Applied for JD Title", 
                    options=jd_options,
                    index=jd_default_index,
                    key=f"jd_select_{resume_name}_{idx}",
                )
                
            with col_date_input:
                try:
                    date_obj = date.fromisoformat(current_submitted_date)
                except (ValueError, TypeError):
                    date_obj = date.today()
                    
                new_submitted_date = st.date_input(
                    "Submitted Date", 
                    value=date_obj,
                    key=f"date_input_{resume_name}_{idx}"
                )
            
            st.markdown("---")
            
            # Dedicated Approve/Reject/Pending Buttons for Quick Actions
            col_quick_approve, col_quick_reject, col_quick_pending, _ = st.columns([1, 1, 1, 5])
            
            jd_to_save = new_applied_jd if new_applied_jd != "Select JD" else "N/A (Pending Assignment)"
            date_to_save = new_submitted_date.strftime("%Y-%m-%d")

            # Function to run status update and RERUN
            def run_update_and_rerun(status_to_set):
                update_resume_metadata(
                    resume_name, 
                    status_to_set, 
                    jd_to_save, 
                    date_to_save,
                    idx
                )
                st.rerun()

            with col_quick_approve:
                # Approve button
                if st.button("✅ Approve", key=f"quick_approve_{resume_name}_{idx}", use_container_width=True):
                    run_update_and_rerun("Approved")

            with col_quick_reject:
                # Reject button
                if st.button("❌ Reject", key=f"quick_reject_{resume_name}_{idx}", use_container_width=True):
                    run_update_and_rerun("Rejected")

            with col_quick_pending:
                 # Pending button
                if st.button("🟡 Pending", key=f"quick_pending_{resume_name}_{idx}", use_container_width=True):
                    run_update_and_rerun("Pending")

            
    st.markdown("---")
            
    # --- Summary of All Resumes (Updated to reflect latest status) ---
    summary_data = []
    for resume_data in st.session_state.resumes_to_analyze:
        name = resume_data['name']
        summary_data.append({
            "Resume": name, 
            "Applied JD": resume_data.get('applied_jd', 'N/A'),
            "Submitted Date": resume_data.get('submitted_date', 'N/A'),
            "Status": st.session_state.resume_statuses.get(name, "Pending")
        })
        
    st.subheader("Summary of All Resumes")
    st.dataframe(summary_data, use_container_width=True)


def vendor_approval_tab_content():
    st.header("🤝 Vendor Approval") 
    
    st.markdown("### 1. Add New Vendor")
    if "vendors" not in st.session_state:
        st.session_state.vendors = []
    if "vendor_statuses" not in st.session_state:
        st.session_state.vendor_statuses = {}
        
    # --- START of Vendor Submission Form ---
    # Using clear_on_submit=True to clear input widgets after successful submission.
    form_key = "add_vendor_form"
    with st.form(form_key, clear_on_submit=True): 
        st.markdown("#### Vendor Company Details")
        col1, col2 = st.columns(2)
        with col1:
            # Note: For form clearing to work, use unique keys inside the form
            vendor_name = st.text_input("Vendor Company Name", key="new_vendor_name_input", help="The legal name of the vendor company.")
        with col2:
            vendor_domain = st.text_input("Service / Domain Name", key="new_vendor_domain_input", help="E.g., HR Consulting, SaaS Platform, Recruitment Agency.")
            
        vendor_code = st.text_input("Vendor ID / Code (if applicable)", key="new_vendor_code_input", help="Internal tracking code or system ID.")
        
        st.markdown("#### Contact & Address Details")
        col3, col4, col5 = st.columns(3)
        with col3:
            contact_person = st.text_input("Contact Person", key="new_contact_person_input")
        with col4:
            contact_email = st.text_input("Email ID", key="new_contact_email_input")
        with col5:
            contact_number = st.text_input("Contact Number", key="new_contact_number_input")
            
        company_address = st.text_area("Company Address", key="new_company_address_input", height=50)

        st.markdown("#### Submission Details")
        col6, col7 = st.columns(2)
        with col6:
            submitted_date = st.date_input("Submitted Date", value=date.today(), key="new_vendor_date_input")
        with col7:
            initial_status = st.selectbox(
                "Set Status", 
                ["Pending Review", "Approved", "Rejected"],
                index=0, 
                key="new_vendor_status_select"
            )
        
        add_vendor_button = st.form_submit_button("Add Vendor", use_container_width=True)
        
        # Use a temporary flag to track if a *new* vendor was successfully added within this block
        st.session_state['vendor_added_flag'] = False 

        if add_vendor_button:
            if vendor_name and contact_person and contact_email:
                vendor_id = vendor_name.strip()
                
                if vendor_id in st.session_state.vendor_statuses:
                    st.warning(f"Vendor '{vendor_name}' already exists.")
                else:
                    new_vendor = {
                        'name': vendor_name.strip(),
                        'domain': vendor_domain.strip(),
                        'code': vendor_code.strip() if vendor_code else 'N/A',
                        'contact_person': contact_person.strip(),
                        'email': contact_email.strip(),
                        'phone': contact_number.strip() if contact_number else 'N/A',
                        'address': company_address.strip() if company_address else 'N/A',
                        'submitted_date': submitted_date.strftime("%Y-%m-%d")
                    }
                    st.session_state.vendors.append(new_vendor)
                    st.session_state.vendor_statuses[vendor_id] = initial_status
                    st.success(f"Vendor **{vendor_name}** added successfully with status **{initial_status}**. Fields are now clear for the next entry.")
                    st.session_state['vendor_added_flag'] = True # Set flag

            else:
                st.error("Please fill in **Vendor Company Name**, **Contact Person**, and **Email ID**.")
                
    # Rerun the page script explicitly outside the form only if new data was added.
    # This updates the Summary table immediately after clearing the form.
    if st.session_state.get('vendor_added_flag'):
         del st.session_state['vendor_added_flag'] # Clear the flag
         st.rerun()

    st.markdown("---")
    
    st.markdown("### 2. Update Existing Vendor Status")
    
    if not st.session_state.vendors:
        st.info("No vendors have been added yet.")
    else:
        # Loop through vendors to display and allow status update
        for idx, vendor in enumerate(st.session_state.vendors):
            vendor_name = vendor['name']
            vendor_id = vendor_name 
            current_status = st.session_state.vendor_statuses.get(vendor_id, "Unknown")
            
            with st.container(border=True):
                
                # --- Display Vendor Info ---
                st.markdown(f"### **{vendor_name}** (Code: `{vendor['code']}`) - **Current Status:** **{current_status}**")
                
                col_domain, col_contact_info = st.columns(2)
                
                with col_domain:
                    st.markdown(f"**Domain:** {vendor['domain']}")
                    st.markdown(f"**Address:** *{vendor['address'].replace('\n', ', ')}*")
                    st.markdown(f"**Submitted:** {vendor['submitted_date']}")
                    
                with col_contact_info:
                    st.markdown(f"**Contact Person:** {vendor['contact_person']}")
                    st.markdown(f"**Email:** `{vendor['email']}`")
                    st.markdown(f"**Phone:** `{vendor['phone']}`")
                    
                st.markdown("---")
                
                # --- Status Update Controls ---
                col_status_input, col_update_btn = st.columns([3, 1])
                
                with col_status_input:
                    new_status = st.selectbox(
                        "Set New Status",
                        ["Pending Review", "Approved", "Rejected"],
                        index=["Pending Review", "Approved", "Rejected"].index(current_status),
                        key=f"vendor_status_select_{idx}",
                    )

                with col_update_btn:
                    st.markdown("##") # Space out button
                    if st.button("Update Status", key=f"vendor_update_btn_{idx}", use_container_width=True):
                        
                        st.session_state.vendor_statuses[vendor_id] = new_status
                        
                        st.success(f"Status for **{vendor_name}** updated to **{new_status}**.")
                        st.rerun()
                        
        st.markdown("---")
        
        summary_data = []
        for vendor in st.session_state.vendors:
            name = vendor['name']
            summary_data.append({
                "Vendor Name": name,
                "Vendor ID / Code": vendor['code'],
                "Domain": vendor['domain'],
                "Contact Person": vendor['contact_person'],
                "Email ID": vendor['email'],
                "Submitted Date": vendor['submitted_date'],
                "Status": st.session_state.vendor_statuses.get(name, "Unknown")
            })
        
        st.subheader("Summary of All Vendors")
        st.dataframe(summary_data, use_container_width=True)


def admin_dashboard():
    st.title("🧑‍💼 Admin Dashboard")
    
    nav_col, _ = st.columns([1, 1]) 

    with nav_col:
        if st.button("🚪 Log Out", use_container_width=True):
            go_to("login") 
    
    # Initialize Admin session state variables (Defensive check)
    if "admin_jd_list" not in st.session_state: st.session_state.admin_jd_list = []
    if "resumes_to_analyze" not in st.session_state: st.session_state.resumes_to_analyze = []
    if "admin_match_results" not in st.session_state: st.session_state.admin_match_results = []
    if "resume_statuses" not in st.session_state: st.session_state.resume_statuses = {}
    if "vendors" not in st.session_state: st.session_state.vendors = []
    if "vendor_statuses" not in st.session_state: st.session_state.vendor_statuses = {}
        
    
    # --- TAB ORDER ---
    tab_jd, tab_analysis, tab_user_mgmt, tab_statistics = st.tabs([
        "📄 JD Management", 
        "📊 Resume Analysis", 
        "🛠️ User Management", 
        "📈 Statistics" 
    ])
    # -------------------------

    # --- TAB 1: JD Management ---
    with tab_jd:
        st.subheader("Add and Manage Job Descriptions (JD)")
        
        jd_type = st.radio("Select JD Type", ["Single JD", "Multiple JD"], key="jd_type_admin")
        st.markdown("### Add JD by:")
        
        method = st.radio("Choose Method", ["Upload File", "Paste Text", "LinkedIn URL"], key="jd_add_method_admin") 

        # URL
        if method == "LinkedIn URL":
            url_list = st.text_area(
                "Enter one or more URLs (comma separated)" if jd_type == "Multiple JD" else "Enter URL", key="url_list_admin"
            )
            if st.button("Add JD(s) from URL", key="add_jd_url_btn_admin"):
                if url_list:
                    urls = [u.strip() for u in url_list.split(",")] if jd_type == "Multiple JD" else [url_list.strip()]
                    
                    count = 0
                    for url in urls:
                        if not url: continue
                        
                        with st.spinner(f"Attempting JD extraction for: {url}"):
                            jd_text = extract_jd_from_linkedin_url(url)
                            metadata = extract_jd_metadata(jd_text) 
                        
                        name_base = url.split('/jobs/view/')[-1].split('/')[0] if '/jobs/view/' in url else f"URL {count+1}"
                        st.session_state.admin_jd_list.append({"name": f"JD from URL: {name_base}", "content": jd_text, **metadata}) 
                        if not jd_text.startswith("[Error"):
                            count += 1
                            
                    if count > 0:
                        st.success(f"✅ {count} JD(s) added successfully! Check the display below for the extracted content.")
                    else:
                        st.error("No JDs were added successfully.")


        # Paste Text
        elif method == "Paste Text":
            text_list = st.text_area(
                "Paste one or more JD texts (separate by '---')" if jd_type == "Multiple JD" else "Paste JD text here", key="text_list_admin"
            )
            if st.button("Add JD(s) from Text", key="add_jd_text_btn_admin"):
                if text_list:
                    texts = [t.strip() for t in text_list.split("---")] if jd_type == "Multiple JD" else [text_list.strip()]
                    for i, text in enumerate(texts):
                         if text:
                            name_base = text.splitlines()[0].strip()
                            if len(name_base) > 30: name_base = f"{name_base[:27]}..."
                            if not name_base: name_base = f"Pasted JD {len(st.session_state.admin_jd_list) + i + 1}"
                            
                            metadata = extract_jd_metadata(text)
                            st.session_state.admin_jd_list.append({"name": name_base, "content": text, **metadata}) 
                    st.success(f"✅ {len(texts)} JD(s) added successfully!")

        # Upload File
        elif method == "Upload File":
            uploaded_files = st.file_uploader(
                "Upload JD file(s)",
                type=["pdf", "txt", "docx"],
                accept_multiple_files=(jd_type == "Multiple JD"),
                key="jd_file_uploader_admin"
            )
            
            if st.button("Add JD(s) from File", key="add_jd_file_btn_admin"):
                if uploaded_files is None:
                    st.warning("Please upload file(s).")
                    
                files_to_process = uploaded_files if isinstance(uploaded_files, list) else ([uploaded_files] if uploaded_files else [])
                
                count = 0
                for file in files_to_process:
                    if file: 
                        temp_dir = tempfile.mkdtemp()
                        temp_path = os.path.join(temp_dir, file.name)
                        with open(temp_path, "wb") as f:
                            f.write(file.getbuffer())
                            
                        file_type = get_file_type(temp_path)
                        jd_text = extract_content(file_type, temp_path)
                        
                        if not jd_text.startswith("Error"):
                            metadata = extract_jd_metadata(jd_text)
                            st.session_state.admin_jd_list.append({"name": file.name, "content": jd_text, **metadata}) 
                            count += 1
                        else:
                            st.error(f"Error extracting content from {file.name}: {jd_text}")
                            
                if count > 0:
                    st.success(f"✅ {count} JD(s) added successfully!")
                elif uploaded_files:
                    st.error("No valid JD files were uploaded or content extraction failed.")


        # Display Added JDs
        if st.session_state.admin_jd_list:
            
            col_display_header, col_clear_button = st.columns([3, 1])
            
            with col_display_header:
                st.markdown("### ✅ Current JDs Added:")
                
            with col_clear_button:
                if st.button("🗑️ Clear All JDs", key="clear_jds_admin", use_container_width=True, help="Removes all currently loaded JDs."):
                    st.session_state.admin_jd_list = []
                    st.session_state.admin_match_results = [] 
                    st.success("All JDs and associated match results have been cleared.")
                    st.rerun() 

            for idx, jd_item in enumerate(st.session_state.admin_jd_list, 1):
                title = jd_item['name']
                display_title = title.replace("--- Simulated JD for: ", "")
                with st.expander(f"JD {idx}: {display_title} | Role: {jd_item.get('role', 'N/A')}"):
                    st.markdown(f"**Job Type:** {jd_item.get('job_type', 'N/A')} | **Key Skills:** {', '.join(jd_item.get('key_skills', ['N/A']))}")
                    st.markdown("---")
                    st.text(jd_item['content'])
        else:
            st.info("No Job Descriptions added yet.")


    # --- TAB 2: Resume Analysis --- 
    with tab_analysis:
        st.subheader("Analyze Resumes Against Job Descriptions")

        # 1. Resume Upload
        st.markdown("#### 1. Upload Resumes")
        
        resume_upload_type = st.radio("Upload Type", ["Single Resume", "Multiple Resumes"], key="resume_upload_type_admin")

        uploaded_files = st.file_uploader(
            "Choose files to analyze",
            type=["pdf", "docx", "txt", "json", "rtf"], 
            accept_multiple_files=(resume_upload_type == "Multiple Resumes"),
            key="resume_file_uploader_admin"
        )
        
        col_parse, col_clear = st.columns([3, 1])
        
        with col_parse:
            if st.button("Load and Parse Resume(s) for Analysis", key="parse_resumes_admin", use_container_width=True):
                if uploaded_files:
                    files_to_process = uploaded_files if isinstance(uploaded_files, list) else ([uploaded_files] if uploaded_files else [])
                    
                    count = 0
                    with st.spinner("Parsing resume(s)... This may take a moment."):
                        for file in files_to_process:
                            if file: 
                                result = parse_and_store_resume(file, file_name_key='admin_analysis', source_type='file')
                                
                                if "error" not in result:
                                    result['applied_jd'] = "N/A (Pending Assignment)"
                                    result['submitted_date'] = date.today().strftime("%Y-%m-%d")
                                    
                                    st.session_state.resumes_to_analyze.append(result)
                                    
                                    resume_id = result['name']
                                    if resume_id not in st.session_state.resume_statuses:
                                        st.session_state.resume_statuses[resume_id] = "Pending"
                                    
                                    count += 1
                                else:
                                    st.error(f"Failed to parse {file.name}: {result['error']}")

                    if count > 0:
                        st.success(f"Successfully loaded and parsed {count} resume(s) for analysis.")
                        st.rerun() 
                    elif not st.session_state.resumes_to_analyze:
                        st.warning("No resumes were successfully loaded and parsed.")
                else:
                    st.warning("Please upload one or more resume files.")
        
        with col_clear:
            if st.button("🗑️ Clear All Resumes", key="clear_resumes_admin", use_container_width=True, help="Removes all currently loaded resumes and match results."):
                st.session_state.resumes_to_analyze = []
                st.session_state.admin_match_results = []
                st.session_state.resume_statuses = {} 
                st.success("All resumes and associated match results have been cleared.")
                st.rerun() 


        st.markdown("---")

        # 2. JD Selection and Analysis
        st.markdown("#### 2. Select JD and Run Analysis")

        if not st.session_state.resumes_to_analyze:
            st.info("Upload and parse resumes first to enable analysis.")
            if not st.session_state.admin_jd_list: return
        elif not st.session_state.admin_jd_list:
            st.error("Please add at least one Job Description in the 'JD Management' tab before running an analysis.")
            return

        resume_names = [r['name'] for r in st.session_state.resumes_to_analyze]
        selected_resume_names = st.multiselect(
            "Select Resume(s) for Matching",
            options=resume_names,
            default=resume_names, 
            key="select_resumes_admin"
        )
        
        resumes_to_match = [
            r for r in st.session_state.resumes_to_analyze 
            if r['name'] in selected_resume_names
        ]

        jd_options = {item['name']: item['content'] for item in st.session_state.admin_jd_list}
        selected_jd_name = st.selectbox("Select JD for Matching", list(jd_options.keys()), key="select_jd_admin")
        selected_jd_content = jd_options.get(selected_jd_name, "")


        if st.button(f"Run Match Analysis on {len(resumes_to_match)} Selected Resume(s)", key="run_match_analysis_admin"):
            st.session_state.admin_match_results = []
            
            if not selected_jd_content:
                st.error("Selected JD content is empty.")
                return
            
            if not resumes_to_match:
                st.warning("No resumes were selected for matching.")
                return

            results_with_score = []
            
            with st.spinner(f"Matching {len(resumes_to_match)} resumes against '{selected_jd_name}'..."):
                for resume_data in resumes_to_match: 
                    
                    resume_name = resume_data['name']
                    parsed_json = resume_data['parsed']

                    try:
                        fit_output = evaluate_jd_fit(selected_jd_content, parsed_json)
                        
                        overall_score_match = re.search(r'Overall Fit Score:\s*[^\d]*(\d+)\s*/10', fit_output, re.IGNORECASE)
                        section_analysis_match = re.search(
                             r'--- Section Match Analysis ---\s*(.*?)\s*Strengths/Matches:', 
                             fit_output, re.DOTALL
                        )

                        skills_percent, experience_percent, education_percent = 'N/A', 'N/A', 'N/A'
                        
                        if section_analysis_match:
                            section_text = section_analysis_match.group(1)
                            skills_match = re.search(r'Skills Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                            experience_match = re.search(r'Experience Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                            education_match = re.search(r'Education Match:\s*\[?(\d+)%\]?', section_text, re.IGNORECASE)
                            
                            if skills_match: skills_percent = skills_match.group(1)
                            if experience_match: experience_percent = experience_match.group(1)
                            if education_match: education_percent = education_match.group(1)
                        
                        overall_score = overall_score_match.group(1) if overall_score_match else 'N/A'
                        
                        results_with_score.append({
                            "resume_name": resume_name,
                            "jd_name": selected_jd_name,
                            "overall_score": overall_score,
                            "numeric_score": int(overall_score) if overall_score.isdigit() else -1, # For sorting
                            "skills_percent": skills_percent,
                            "experience_percent": experience_percent, 
                            "education_percent": education_percent,   
                            "full_analysis": fit_output
                        })
                    except Exception as e:
                        results_with_score.append({
                            "resume_name": resume_name,
                            "jd_name": selected_jd_name,
                            "overall_score": "Error",
                            "numeric_score": -1, # Set a low score for errors
                            "skills_percent": "Error",
                            "experience_percent": "Error", 
                            "education_percent": "Error",   
                            "full_analysis": f"Error running analysis: {e}\n{traceback.format_exc()}"
                        })
                
                results_with_score.sort(key=lambda x: x['numeric_score'], reverse=True)
                st.session_state.admin_match_results = results_with_score

            st.success("Analysis complete!")


        # 3. Display Results
        if st.session_state.get('admin_match_results'):
            st.markdown("#### 3. Match Results")
            results_df = st.session_state.admin_match_results
            
            display_data = []
            for item in results_df:
                
                display_data.append({
                    "Resume": item["resume_name"],
                    "JD": item["jd_name"],
                    "Fit Score (out of 10)": item["overall_score"],
                    "Skills (%)": item.get("skills_percent", "N/A"),
                    "Experience (%)": item.get("experience_percent", "N/A"), 
                    "Education (%)": item.get("education_percent", "N/A"),
                })

            st.dataframe(display_data, use_container_width=True)

            st.markdown("##### Detailed Reports")
            for item in results_df:
                status = st.session_state.resume_statuses.get(item["resume_name"], 'Pending') 
                header_text = f"Report for **{item['resume_name']}** against {item['jd_name']} (Score: **{item['overall_score']}/10** | S: **{item.get('skills_percent', 'N/A')}%** | E: **{item.get('experience_percent', 'N/A')}%** | Edu: **{item.get('education_percent', 'N/A')}%**) - Current Status: {status}"
                with st.expander(header_text):
                    st.markdown(item['full_analysis'])

                    
    # --- TAB 3: User Management (Parent Tab) ---
    with tab_user_mgmt:
        st.header("🛠️ User Management")
        
        nested_tab_candidate, nested_tab_vendor = st.tabs([
            "👤 Candidate Approval",
            "🤝 Vendor Approval"
        ])
        
        with nested_tab_candidate:
            candidate_approval_tab_content() 
            
        with nested_tab_vendor:
            vendor_approval_tab_content() 
            

    # --- TAB 4: Statistics (UPDATED) ---
    with tab_statistics:
        st.header("System Statistics")
        st.markdown("---")

        total_candidates = len(st.session_state.resumes_to_analyze)
        total_jds = len(st.session_state.admin_jd_list)
        total_vendors = len(st.session_state.vendors)
        no_of_applications = total_candidates 
        
        # --- Top-Level Metrics ---
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(label="Total Candidates", value=total_candidates, delta="Resumes Submitted")

        with col2:
            st.metric(label="Total JDs", value=total_jds, delta_color="off")
        
        with col3:
            st.metric(label="Total Vendors", value=total_vendors, delta_color="off")

        with col4:
            # We can use this to show total resumes, as per the code logic
            st.metric(label="Total Applications", value=no_of_applications, delta_color="off")
            
        st.markdown("---")
        
        # --- Candidate Status Breakdown ---
        st.subheader("Candidate Status Breakdown (Resumes)")
        
        candidate_status_counts = {}
        for status in st.session_state.resume_statuses.values():
            # Standardizing status names for display
            display_status = status.replace(' ', '') 
            candidate_status_counts[display_status] = candidate_status_counts.get(display_status, 0) + 1
            
        status_cols_cand = st.columns(max(len(candidate_status_counts), 1))
        
        if candidate_status_counts:
            # Display metrics in columns, ensuring maximum 3 per row for good spacing
            for i, (status, count) in enumerate(candidate_status_counts.items()):
                with status_cols_cand[i % len(status_cols_cand)]:
                    # Display status in a user-friendly format
                    display_label = status.title().replace('Of', ' of').replace('Awaiting', 'Awaiting')
                    st.metric(label=f"Candidates {display_label}", value=count)
        else:
            st.info("No resumes loaded to calculate status breakdown.")

        st.markdown("---")
        
        # --- Vendor Status Breakdown (NEW) ---
        st.subheader("Vendor Status Breakdown")
        
        vendor_status_counts = {}
        for status in st.session_state.vendor_statuses.values():
            # Standardizing status names for display
            display_status = status.replace(' ', '')
            vendor_status_counts[display_status] = vendor_status_counts.get(display_status, 0) + 1
            
        status_cols_vend = st.columns(max(len(vendor_status_counts), 1))
        
        if vendor_status_counts:
            # Display metrics in columns
            for i, (status, count) in enumerate(vendor_status_counts.items()):
                with status_cols_vend[i % len(status_cols_vend)]:
                    display_label = status.title().replace('Of', ' of')
                    st.metric(label=f"Vendors {display_label}", value=count)
        else:
            st.info("No vendors added to calculate status breakdown.")


# --- Session State & Main Function Initialization (Required for execution) ---
if __name__ == '__main__':
    st.set_page_config(layout="wide", page_title="PragyanAI Admin Dashboard")
    # Initialize state for necessary variables
    if 'page' not in st.session_state: st.session_state.page = "admin_dashboard" 
    if 'admin_jd_list' not in st.session_state: st.session_state.admin_jd_list = []
    if 'resumes_to_analyze' not in st.session_state: st.session_state.resumes_to_analyze = []
    if 'admin_match_results' not in st.session_state: st.session_state.admin_match_results = []
    if 'resume_statuses' not in st.session_state: st.session_state.resume_statuses = {}
    if 'vendors' not in st.session_state: st.session_state.vendors = []
    if 'vendor_statuses' not in st.session_state: st.session_state.vendor_statuses = {}
    
    if st.session_state.page == "admin_dashboard":
        admin_dashboard()
    else:
        st.info("Log in as Admin to view dashboard.")
